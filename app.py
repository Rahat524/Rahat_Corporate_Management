
from __future__ import annotations
from flask import Flask, jsonify, request, send_file, send_from_directory, session, g
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
import sqlite3, json, io, os, socket, shutil, tempfile, zipfile
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

BASE = Path(__file__).resolve().parent
SEED = BASE / "data" / "seed_data.json"
UPLOADS = BASE / "uploads"
UPLOADS.mkdir(exist_ok=True)

# Keep business data outside the extracted version folder. Every V6/V7/V8 update
# therefore uses the same database automatically.
APP_DATA = Path(os.environ.get("LOCALAPPDATA") or (Path.home() / ".rahat_corporate_management")) / "RahatCorporateManagement"
APP_DATA.mkdir(parents=True, exist_ok=True)
DB = Path(os.environ.get("RAHAT_DATABASE_PATH") or os.environ.get("DATABASE_PATH") or (APP_DATA / "corporate_scrap.db"))
DB.parent.mkdir(parents=True, exist_ok=True)
BACKUP_DIR = APP_DATA / "Backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
PROJECT_DB = BASE / "data" / "corporate_scrap.db"
USERS_BACKUP = APP_DATA / "users_permanent_backup.json"
PROJECT_USERS_BACKUP = BASE / "data" / "users_permanent_backup.json"

STORES = {
    "S012": "ISB-MEGA-RAMZAN MALL",
    "S013": "ISB-MEGA-GULBERG GREEN",
    "S014": "LHR - MEGA - GULBERG",
    "S015": "LHR - SUP - HALY TOWER",
    "S016": "FSD-MEGA-MISAQ MALL",
    "S017": "GUJ- MEGA - KINGS MALL",
    "S018": "SKT - MEGA - HARRAR",
    "S019": "BWP – MEGA – SS TOWER",
    "S020": "GJT - MEGA - GUJRAT",
    "S022": "SGD- MEGA - SARGODHA",
    "S023": "VHR-MEGA - VEHARI",
    "S024": "LHR - MEGA - BAHRIA TOWN",
    "S025": "LHR - MEGA - L3",
    "S026": "PSH - MEGA - PESHAWAR",
    "S028": "ISB - MEGA - BARA KAHU",
    "S030": "LHR-MEGA-PS MALL L5",
    "S031": "MLT-MEGA-BUCH VILLAS",
    "S033": "SWL - MEGA - SAHIWAL",
    "S034": "ISB E-11",
}
DEFAULT_STORE = "S024"
STORE_DATA_DIR = APP_DATA / "Stores"
STORE_DATA_DIR.mkdir(parents=True, exist_ok=True)

def auth_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def active_store_code():
    try:
        user = getattr(g, "user", None) or {}
        access = str(user.get("store_access") or "ALL").strip().upper()
        if access != "ALL" and access in STORES:
            return access
        selected = str(session.get("active_store") or DEFAULT_STORE).strip().upper()
        return selected if selected in STORES else DEFAULT_STORE
    except RuntimeError:
        return DEFAULT_STORE

def store_db_path(code=None):
    code = (code or active_store_code()).upper()
    if code == DEFAULT_STORE:
        return DB
    return STORE_DATA_DIR / f"{code}.db"

def initialize_store_databases():
    """Create isolated store databases and apply one-time S024-only Cashier Closing cleanup."""
    for code in STORES:
        if code == DEFAULT_STORE:
            continue
        target = store_db_path(code)
        is_new = not target.exists() or target.stat().st_size == 0
        if is_new:
            shutil.copy2(DB, target)

        conn = sqlite3.connect(target)
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS app_migrations(migration_key TEXT PRIMARY KEY, applied_at TEXT)")

            # v4 is intentionally a new marker. Older online databases may already
            # contain the v3 marker even though copied S024 cashier rows remained.
            # Run once per non-S024 store, clearing ONLY Cashier Closing data.
            migration_key = "cashier_closing_s024_only_v4_20260807"
            already = conn.execute(
                "SELECT 1 FROM app_migrations WHERE migration_key=?", (migration_key,)
            ).fetchone()
            if not already:
                for table in ("cashier_closings", "cashier_employees"):
                    try:
                        conn.execute(f"DELETE FROM {table}")
                    except sqlite3.Error:
                        pass
                conn.execute(
                    "INSERT OR REPLACE INTO app_migrations(migration_key,applied_at) VALUES(?,?)",
                    (migration_key, datetime.now().isoformat(timespec="seconds")),
                )
                conn.commit()
        finally:
            conn.close()

def backup_users(conn=None):
    """Keep a second copy of user accounts outside the update folder."""
    own = conn is None
    conn = conn or auth_db()
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("SELECT full_name,username,password_hash,role_name,user_type,store_access,permissions,status,created_at,last_login FROM users").fetchall()
        payload = [dict(r) for r in rows]
        text = json.dumps(payload, ensure_ascii=False, indent=2)
        for target in (USERS_BACKUP, PROJECT_USERS_BACKUP):
            try:
                target.parent.mkdir(parents=True, exist_ok=True)
                tmp = target.with_suffix(target.suffix + ".tmp")
                tmp.write_text(text, encoding="utf-8")
                os.replace(tmp, target)
            except OSError:
                pass
    except sqlite3.Error:
        pass
    finally:
        if own:
            conn.close()

def restore_users_from_backup(conn):
    """Restore/merge accounts from all available permanent backup copies."""
    merged = {}
    for source in (PROJECT_USERS_BACKUP, USERS_BACKUP):
        if not source.exists():
            continue
        try:
            payload = json.loads(source.read_text(encoding="utf-8"))
        except Exception:
            continue
        for u in payload if isinstance(payload, list) else []:
            username = str(u.get("username") or "").strip()
            if username:
                merged[username.lower()] = u
    for u in merged.values():
        username = str(u.get("username") or "").strip()
        conn.execute("""INSERT OR IGNORE INTO users(
            full_name,username,password_hash,role_name,user_type,store_access,permissions,status,created_at,last_login
        ) VALUES(?,?,?,?,?,?,?,?,?,?)""", (
            u.get("full_name") or username, username, u.get("password_hash") or "",
            u.get("role_name") or "Master Account", u.get("user_type") or "Local",
            u.get("store_access") or "ALL", u.get("permissions") or "[]",
            u.get("status") or "Active", u.get("created_at") or datetime.now().isoformat(timespec="seconds"),
            u.get("last_login")
        ))

def _db_record_score(path: Path):
    try:
        conn = sqlite3.connect(path)
        score = 0
        for table in ("customer_ledger", "vendor_ledger", "cash_ledger", "deleted_cash_entries", "customers", "vendors", "users"):
            try:
                score += int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])
            except sqlite3.Error:
                pass
        conn.close()
        return score
    except Exception:
        return -1

def discover_and_migrate_legacy_database():
    """Migrate legacy data only on first run; never overwrite an existing permanent database."""
    # Once the permanent database exists, it is the source of truth. Updating or
    # replacing the application folder must never replace users, passwords,
    # permissions, or business entries stored in LocalAppData.
    if DB.exists() and DB.stat().st_size > 0:
        return
    candidates = []
    roots = [BASE.parent, BASE.parent.parent, Path.home() / "Desktop"]
    seen = set()
    for root in roots:
        try:
            if not root.exists():
                continue
            for candidate in root.glob("**/data/corporate_scrap.db"):
                try:
                    resolved = candidate.resolve()
                    if resolved in seen or resolved == DB.resolve() or ".venv" in resolved.parts:
                        continue
                    seen.add(resolved)
                    candidates.append(candidate)
                except Exception:
                    continue
        except Exception:
            continue
    if PROJECT_DB.exists() and PROJECT_DB not in candidates:
        candidates.append(PROJECT_DB)
    valid = [(_db_record_score(p), p.stat().st_mtime, p) for p in candidates if p.exists()]
    valid = [x for x in valid if x[0] >= 0]
    current_score = _db_record_score(DB) if DB.exists() else -1
    if valid:
        best_score, _, best = max(valid, key=lambda x: (x[0], x[1]))
        if best_score > current_score:
            if DB.exists() and DB.stat().st_size > 0:
                safety = BACKUP_DIR / f"before_auto_migration_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                shutil.copy2(DB, safety)
            shutil.copy2(best, DB)
            print(f"Existing data migrated automatically from: {best}")

def create_automatic_backup(force=False):
    if not DB.exists():
        return None
    today = datetime.now().strftime("%Y-%m-%d")
    target = BACKUP_DIR / f"corporate_scrap_auto_{today}.db"
    if force or not target.exists():
        tmp = target.with_suffix(".tmp")
        shutil.copy2(DB, tmp)
        os.replace(tmp, target)
    # Keep the latest 30 automatic backups.
    backups = sorted(BACKUP_DIR.glob("corporate_scrap_auto_*.db"), key=lambda x: x.stat().st_mtime, reverse=True)
    for old in backups[30:]:
        try: old.unlink()
        except OSError: pass
    return target

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.environ.get("APP_SECRET_KEY", "change-this-secret-in-production")
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax")

@app.after_request
def disable_api_cache(response):
    """Always show the latest shared records to every logged-in user."""
    if request.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

PERMISSIONS = [
    'dashboard_view',
    'dashboard_add',
    'dashboard_edit',
    'dashboard_delete',
    'dashboard_print',
    'dashboard_export',
    'corporate_customers_view',
    'corporate_customers_add',
    'corporate_customers_edit',
    'corporate_customers_delete',
    'corporate_customers_print',
    'corporate_customers_export',
    'vendor_management_view',
    'vendor_management_add',
    'vendor_management_edit',
    'vendor_management_delete',
    'vendor_management_print',
    'vendor_management_export',
    'head_cash_view',
    'head_cash_add',
    'head_cash_edit',
    'head_cash_delete',
    'head_cash_print',
    'head_cash_export',
    'petty_cash_view',
    'petty_cash_add',
    'petty_cash_edit',
    'petty_cash_delete',
    'petty_cash_print',
    'petty_cash_export',
    'cashier_closing_view',
    'cashier_closing_add',
    'cashier_closing_edit',
    'cashier_closing_delete',
    'cashier_closing_print',
    'cashier_closing_export',
    'return_counter_view',
    'return_counter_add',
    'return_counter_edit',
    'return_counter_delete',
    'return_counter_print',
    'return_counter_export',
    'lost_found_view',
    'lost_found_add',
    'lost_found_edit',
    'lost_found_delete',
    'lost_found_print',
    'lost_found_export',
    'theft_view',
    'theft_add',
    'theft_edit',
    'theft_delete',
    'theft_print',
    'theft_export',
    'documents_data_view',
    'documents_data_add',
    'documents_data_edit',
    'documents_data_delete',
    'documents_data_print',
    'documents_data_export',
    'reports_view',
    'reports_add',
    'reports_edit',
    'reports_delete',
    'reports_print',
    'reports_export',
    'security_users_view',
    'security_users_add',
    'security_users_edit',
    'security_users_delete',
    'security_users_print',
    'security_users_export'
]

def db():
    conn = sqlite3.connect(store_db_path())
    conn.row_factory = sqlite3.Row
    return conn

def clean_customer_code(value):
    """Return a stable digit-only customer code (Excel numbers may arrive as 12001680.0)."""
    if value is None:
        return ""
    text = str(value).strip().replace(",", "")
    if text.endswith(".0"):
        text = text[:-2]
    return "".join(ch for ch in text if ch.isdigit())

def master_customer_map_from_seed():
    if not SEED.exists():
        return {}
    try:
        seed = json.loads(SEED.read_text(encoding="utf-8"))
        return {
            clean_customer_code(r.get("Customer Code")): str(r.get("Customer Name", "")).strip()
            for r in seed.get("customer_master", [])
            if clean_customer_code(r.get("Customer Code"))
        }
    except Exception:
        return {}

def find_master_customer_code(row, valid_codes):
    """Find the real corporate customer code without trusting SAP offsetting-account columns."""
    preferred = (
        "Customer Code", "Customer Account", "Customer", "Account",
        "Account Number", "Customer No", "Customer Number", "Business Partner",
        "BP Number", "Sold-to Party", "Payer"
    )
    for alias in preferred:
        code = clean_customer_code(pick(row, alias))
        if code in valid_codes:
            return code
    # SAP exports vary. As a safe fallback, scan the full row and accept only an exact master code.
    for value in row.values():
        code = clean_customer_code(value)
        if code in valid_codes:
            return code
    return ""

def seed_cash_data_from_excel(cur):
    source = BASE / "data" / "HC_PC_Original.xlsx"
    if not source.exists() or cur.execute("SELECT COUNT(*) FROM cash_ledger").fetchone()[0] > 0:
        return
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1,max_row=1))]
    ix = {h:i for i,h in enumerate(headers)}
    batch=[]
    now=datetime.now().isoformat(timespec="seconds")
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        gl_text=str(row[ix.get("G/L Acct Long Text",6)] or "").strip()
        if "Head Cashier" in gl_text: cash_type="Head Cash"
        elif "Petty Cash" in gl_text: cash_type="Petty Cash"
        else: continue
        doc=str(row[ix.get("Document Number",0)] or "").strip()
        doc_date=row[ix.get("Document Date",1)]
        post_date=row[ix.get("Posting Date",2)]
        desc=str(row[ix.get("Text",13)] or "").strip()
        try: amount=float(row[ix.get("Amount in Local Currency",11)] or 0)
        except Exception: amount=0.0
        # Cash ledgers use receipt/payment presentation requested by the user:
        # positive SAP amount is shown as Credit; negative amount is shown as Debit.
        debit=abs(amount) if amount<0 else 0.0
        credit=amount if amount>0 else 0.0
        def d(v): return v.strftime("%Y-%m-%d") if hasattr(v,"strftime") else str(v or "")
        batch.append((cash_type,doc,d(doc_date),d(post_date),desc,debit,credit,row_no,now))
        if len(batch)>=2000:
            cur.executemany("INSERT INTO cash_ledger(cash_type,document_number,document_date,posting_date,description,debit,credit,source_row,created_at) VALUES(?,?,?,?,?,?,?,?,?)",batch); batch=[]
    if batch: cur.executemany("INSERT INTO cash_ledger(cash_type,document_number,document_date,posting_date,description,debit,credit,source_row,created_at) VALUES(?,?,?,?,?,?,?,?,?)",batch)
    wb.close()


def seed_cashier_closing_from_excel(cur):
    source = BASE / "data" / "Cashier Closing Report FTMO Aug-26 Updatd.xlsx"
    if not source.exists() or cur.execute("SELECT COUNT(*) FROM cashier_closings").fetchone()[0] > 0:
        return
    wb = load_workbook(source, read_only=True, data_only=True)
    if "Employee Database" in wb.sheetnames:
        for row in wb["Employee Database"].iter_rows(min_row=2, values_only=True):
            if row and row[0] not in (None, ""):
                cur.execute("INSERT OR REPLACE INTO cashier_employees(employee_id,employee_name) VALUES(?,?)",(str(row[0]).replace('.0',''),str(row[1] or '').strip()))
    now=datetime.now().isoformat(timespec="seconds")
    sql="""INSERT OR IGNORE INTO cashier_closings(
        closing_date,employee_id,employee_name,first_5000,first_1000,first_500,first_total,
        second_5000,second_1000,second_500,second_total,third_5000,third_1000,third_500,third_total,
        fourth_5000,fourth_1000,fourth_500,fourth_total,close_5000,close_1000,close_500,close_100,
        close_75,close_50,close_20,close_10,close_5,close_2,close_1,total_closing_cash,system_total_sale,
        collection_difference,audit_status,remarks,ivend_pos,settlement_bank,card_difference,card_status,card_remarks,
        source_sheet,created_at) VALUES("""+",".join(["?"]*42)+")"
    batch=[]
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        ws=wb[sheet_name]
        raw_date=ws.cell(2,1).value
        if not raw_date:
            continue
        closing_date=raw_date.strftime("%Y-%m-%d") if hasattr(raw_date,"strftime") else str(raw_date)
        for row in ws.iter_rows(min_row=5, max_col=39, values_only=True):
            emp=row[0]
            if emp in (None,"","-"):
                continue
            emp_id=str(emp).replace('.0','')
            name=str(row[1] or '').strip()
            if name == '#N/A': name='Employee Not Found'
            vals=[closing_date,emp_id,name]
            nums=[]
            for v in row[2:32]:
                try: nums.append(float(v or 0))
                except: nums.append(0.0)
            vals += nums
            vals += [str(row[32] or ''),str(row[33] or ''),float(row[34] or 0),float(row[35] or 0),float(row[36] or 0),str(row[37] or ''),str(row[38] or ''),sheet_name,now]
            batch.append(tuple(vals))
            if len(batch)>=500:
                cur.executemany(sql,batch); batch=[]
    if batch: cur.executemany(sql,batch)
    wb.close()

def seed_return_approvers_from_excel(cur):
    source = BASE / "data" / "Return Exchange Report FTMO Aug-26.xlsx"
    if not source.exists() or cur.execute("SELECT COUNT(*) FROM return_approvers").fetchone()[0] > 0:
        return
    wb = load_workbook(source, read_only=True, data_only=True)
    if "Management" in wb.sheetnames:
        ws = wb["Management"]
        for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
            designation, name, limit = row
            if not name: continue
            unlimited = 1 if str(limit or '').strip().lower() == 'unlimited' or 'store manager' in str(designation or '').lower() else 0
            try: amount = float(limit or 0) if not unlimited else 0
            except: amount = 0
            cur.execute("INSERT OR IGNORE INTO return_approvers(designation,management_name,approval_limit,unlimited,active) VALUES(?,?,?,?,1)",(str(designation or ''),str(name).strip(),amount,unlimited))
    wb.close()

def init_db():
    discover_and_migrate_legacy_database()
    conn = db()
    cur = conn.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS customers(
        code TEXT PRIMARY KEY,
        name TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS customer_ledger(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_code TEXT,
        customer_name TEXT,
        posting_date TEXT,
        document_date TEXT,
        document_type TEXT,
        document_number TEXT,
        text TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        net_amount REAL DEFAULT 0,
        running_balance REAL DEFAULT 0,
        currency TEXT,
        clearing_document TEXT,
        gl_account TEXT,
        special_gl TEXT,
        offsetting_type TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_customer_ledger_code ON customer_ledger(customer_code);
    CREATE INDEX IF NOT EXISTS idx_customer_ledger_doc ON customer_ledger(document_number);
    CREATE INDEX IF NOT EXISTS idx_customer_ledger_date ON customer_ledger(posting_date);

    CREATE TABLE IF NOT EXISTS duplicate_documents(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_code TEXT,
        customer_name TEXT,
        posting_date TEXT,
        document_date TEXT,
        document_type TEXT,
        document_number TEXT,
        text TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        net_amount REAL DEFAULT 0,
        currency TEXT,
        clearing_document TEXT,
        gl_account TEXT,
        special_gl TEXT,
        offsetting_type TEXT,
        duplicate_occurrences INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS vendors(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vendor_code TEXT UNIQUE,
        vendor_name TEXT NOT NULL,
        security_deposit REAL DEFAULT 0,
        advance REAL DEFAULT 0,
        phone TEXT,
        description TEXT,
        status TEXT DEFAULT 'Active'
    );
    CREATE TABLE IF NOT EXISTS vendor_ledger(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vendor_code TEXT,
        vendor_name TEXT,
        tx_date TEXT,
        tx_type TEXT,
        description TEXT,
        document_number TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        payment_status TEXT DEFAULT 'Pending',
        gate_pass TEXT,
        bilty_no TEXT,
        created_at TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_vendor_ledger_code ON vendor_ledger(vendor_code);
    CREATE INDEX IF NOT EXISTS idx_vendor_ledger_date ON vendor_ledger(tx_date);

    CREATE TABLE IF NOT EXISTS cash_ledger(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cash_type TEXT NOT NULL,
        document_number TEXT,
        document_date TEXT,
        posting_date TEXT,
        description TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        source_row INTEGER,
        created_at TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_cash_type ON cash_ledger(cash_type);
    CREATE INDEX IF NOT EXISTS idx_cash_doc ON cash_ledger(document_number);
    CREATE INDEX IF NOT EXISTS idx_cash_date ON cash_ledger(posting_date);
    CREATE TABLE IF NOT EXISTS cashier_employees(
        employee_id TEXT PRIMARY KEY,
        employee_name TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS cashier_closings(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        closing_date TEXT NOT NULL,
        employee_id TEXT NOT NULL,
        employee_name TEXT,
        first_5000 REAL DEFAULT 0, first_1000 REAL DEFAULT 0, first_500 REAL DEFAULT 0, first_total REAL DEFAULT 0,
        second_5000 REAL DEFAULT 0, second_1000 REAL DEFAULT 0, second_500 REAL DEFAULT 0, second_total REAL DEFAULT 0,
        third_5000 REAL DEFAULT 0, third_1000 REAL DEFAULT 0, third_500 REAL DEFAULT 0, third_total REAL DEFAULT 0,
        fourth_5000 REAL DEFAULT 0, fourth_1000 REAL DEFAULT 0, fourth_500 REAL DEFAULT 0, fourth_total REAL DEFAULT 0,
        close_5000 REAL DEFAULT 0, close_1000 REAL DEFAULT 0, close_500 REAL DEFAULT 0, close_100 REAL DEFAULT 0,
        close_75 REAL DEFAULT 0, close_50 REAL DEFAULT 0, close_20 REAL DEFAULT 0, close_10 REAL DEFAULT 0,
        close_5 REAL DEFAULT 0, close_2 REAL DEFAULT 0, close_1 REAL DEFAULT 0, total_closing_cash REAL DEFAULT 0,
        system_total_sale REAL DEFAULT 0, collection_difference REAL DEFAULT 0, audit_status TEXT, remarks TEXT,
        ivend_pos REAL DEFAULT 0, settlement_bank REAL DEFAULT 0, card_difference REAL DEFAULT 0, card_status TEXT, card_remarks TEXT,
        source_sheet TEXT, created_at TEXT,
        UNIQUE(closing_date, employee_id)
    );
    CREATE INDEX IF NOT EXISTS idx_cashier_closing_date ON cashier_closings(closing_date);
    CREATE INDEX IF NOT EXISTS idx_cashier_closing_emp ON cashier_closings(employee_id);

    CREATE TABLE IF NOT EXISTS deleted_cash_entries(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        original_id INTEGER,
        cash_type TEXT,
        document_number TEXT,
        document_date TEXT,
        posting_date TEXT,
        description TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        deleted_by TEXT,
        deleted_at TEXT,
        delete_reason TEXT
    );

    CREATE TABLE IF NOT EXISTS return_approvers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        designation TEXT,
        management_name TEXT UNIQUE,
        approval_limit REAL,
        unlimited INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS return_entries(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entry_date TEXT,
        serial_no TEXT,
        customer_name TEXT,
        contact_no TEXT,
        return_trx TEXT,
        trx_no TEXT,
        item_description TEXT,
        item_exp_date TEXT,
        qty REAL DEFAULT 0,
        total_amount REAL DEFAULT 0,
        reason TEXT,
        approval_name TEXT,
        designation TEXT,
        approval_limit REAL DEFAULT 0,
        system_status TEXT,
        cash_voucher TEXT,
        received_by TEXT,
        cctv_time TEXT,
        trx_time TEXT,
        source_file TEXT,
        source_sheet TEXT,
        created_at TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_return_date ON return_entries(entry_date);
    CREATE INDEX IF NOT EXISTS idx_return_trx ON return_entries(trx_no);
    CREATE INDEX IF NOT EXISTS idx_return_approval ON return_entries(approval_name);

    CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_name TEXT NOT NULL,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role_name TEXT DEFAULT 'Local User',
        user_type TEXT DEFAULT 'Local',
        store_access TEXT DEFAULT 'ALL',
        permissions TEXT DEFAULT '[]',
        status TEXT DEFAULT 'Active',
        created_at TEXT,
        last_login TEXT
    );
    CREATE TABLE IF NOT EXISTS audit_log(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        action TEXT,
        module TEXT,
        details TEXT,
        created_at TEXT,
        ip_address TEXT
    );
    CREATE TABLE IF NOT EXISTS app_migrations(
        migration_key TEXT PRIMARY KEY,
        applied_at TEXT
    );
    """)
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS reconciliation_runs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_type TEXT NOT NULL,
        title TEXT,
        created_by TEXT,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS reconciliation_results(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id INTEGER NOT NULL,
        source_name TEXT,
        document_number TEXT,
        source_amount REAL DEFAULT 0,
        target_amount REAL DEFAULT 0,
        difference REAL DEFAULT 0,
        status TEXT,
        details TEXT,
        FOREIGN KEY(run_id) REFERENCES reconciliation_runs(id)
    );
    CREATE INDEX IF NOT EXISTS idx_recon_run ON reconciliation_results(run_id);
    CREATE INDEX IF NOT EXISTS idx_recon_status ON reconciliation_results(status);
    CREATE TABLE IF NOT EXISTS period_locks(
        period_key TEXT PRIMARY KEY,
        locked_by TEXT,
        locked_at TEXT,
        reason TEXT
    );
    CREATE TABLE IF NOT EXISTS approval_requests(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        request_type TEXT,
        reference_no TEXT,
        amount REAL DEFAULT 0,
        reason TEXT,
        maker TEXT,
        checker TEXT,
        approver TEXT,
        status TEXT DEFAULT 'Pending Checker',
        created_at TEXT,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS finance_budgets(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_key TEXT NOT NULL,
        cost_center TEXT DEFAULT 'ALL',
        account_head TEXT NOT NULL,
        budget_amount REAL DEFAULT 0,
        created_by TEXT,
        created_at TEXT,
        UNIQUE(period_key,cost_center,account_head)
    );
    CREATE TABLE IF NOT EXISTS month_end_tasks(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_key TEXT NOT NULL,
        task_name TEXT NOT NULL,
        owner TEXT,
        due_date TEXT,
        status TEXT DEFAULT 'Pending',
        remarks TEXT,
        updated_by TEXT,
        updated_at TEXT,
        UNIQUE(period_key,task_name)
    );
    CREATE TABLE IF NOT EXISTS journal_vouchers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        voucher_no TEXT NOT NULL UNIQUE,
        posting_date TEXT NOT NULL,
        narration TEXT,
        status TEXT DEFAULT 'Posted',
        created_by TEXT,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS journal_lines(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        voucher_id INTEGER NOT NULL,
        account_code TEXT,
        account_name TEXT NOT NULL,
        cost_center TEXT DEFAULT 'ALL',
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        FOREIGN KEY(voucher_id) REFERENCES journal_vouchers(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_journal_date ON journal_vouchers(posting_date);
    CREATE INDEX IF NOT EXISTS idx_journal_account ON journal_lines(account_name);
    CREATE TABLE IF NOT EXISTS accrual_schedules(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        schedule_type TEXT NOT NULL,
        reference_no TEXT,
        description TEXT NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        total_amount REAL DEFAULT 0,
        recognized_amount REAL DEFAULT 0,
        account_name TEXT,
        cost_center TEXT DEFAULT 'ALL',
        status TEXT DEFAULT 'Open',
        created_by TEXT,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS fixed_assets(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        asset_code TEXT NOT NULL UNIQUE,
        asset_name TEXT NOT NULL,
        category TEXT,
        acquisition_date TEXT NOT NULL,
        cost REAL DEFAULT 0,
        salvage_value REAL DEFAULT 0,
        useful_life_months INTEGER DEFAULT 60,
        location TEXT,
        custodian TEXT,
        status TEXT DEFAULT 'Active',
        created_by TEXT,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS tax_register(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        posting_date TEXT NOT NULL,
        tax_type TEXT NOT NULL,
        party_type TEXT DEFAULT 'Vendor',
        party_code TEXT,
        party_name TEXT,
        document_number TEXT,
        taxable_amount REAL DEFAULT 0,
        tax_rate REAL DEFAULT 0,
        tax_amount REAL DEFAULT 0,
        status TEXT DEFAULT 'Pending',
        created_by TEXT,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS payment_schedules(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        due_date TEXT NOT NULL,
        party_type TEXT DEFAULT 'Vendor',
        party_code TEXT,
        party_name TEXT NOT NULL,
        reference_no TEXT,
        amount REAL DEFAULT 0,
        priority TEXT DEFAULT 'Normal',
        payment_method TEXT,
        status TEXT DEFAULT 'Planned',
        remarks TEXT,
        created_by TEXT,
        created_at TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_tax_period ON tax_register(posting_date);
    CREATE INDEX IF NOT EXISTS idx_payment_due ON payment_schedules(due_date);
    """)
    # Schema migration for older permanent databases.
    vendor_columns = {r[1] for r in cur.execute("PRAGMA table_info(vendor_ledger)").fetchall()}
    if "document_number" not in vendor_columns:
        cur.execute("ALTER TABLE vendor_ledger ADD COLUMN document_number TEXT")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_vendor_ledger_doc ON vendor_ledger(document_number)")
    conn.commit()

    # Keep the 18-customer master authoritative on every startup. This also repairs
    # databases created by older versions that imported SAP offsetting accounts.
    seed_master = master_customer_map_from_seed()
    for code, name in seed_master.items():
        cur.execute("INSERT OR REPLACE INTO customers(code,name) VALUES(?,?)", (code, name))
    if seed_master:
        placeholders = ",".join("?" for _ in seed_master)
        valid_codes = tuple(seed_master.keys())
        cur.execute(f"DELETE FROM customer_ledger WHERE customer_code NOT IN ({placeholders})", valid_codes)
        cur.execute(f"DELETE FROM duplicate_documents WHERE customer_code NOT IN ({placeholders})", valid_codes)
        for code, name in seed_master.items():
            cur.execute("UPDATE customer_ledger SET customer_name=? WHERE customer_code=?", (name, code))
            cur.execute("UPDATE duplicate_documents SET customer_name=? WHERE customer_code=?", (name, code))
    conn.commit()

    if cur.execute("SELECT COUNT(*) FROM customer_ledger").fetchone()[0] == 0 and SEED.exists():
        seed = json.loads(SEED.read_text(encoding="utf-8"))
        for r in seed.get("customer_master", []):
            cur.execute("INSERT OR IGNORE INTO customers(code,name) VALUES(?,?)",
                        (str(r.get("Customer Code","")), str(r.get("Customer Name",""))))
        customer_master = {
            clean_customer_code(r.get("Customer Code")): str(r.get("Customer Name", "")).strip()
            for r in seed.get("customer_master", [])
            if clean_customer_code(r.get("Customer Code"))
        }
        for r in seed.get("compiled_ledger", []):
            code = clean_customer_code(r.get("Customer Code"))
            if code not in customer_master:
                continue
            cur.execute("""INSERT INTO customer_ledger(
                customer_code,customer_name,posting_date,document_date,document_type,
                document_number,text,debit,credit,net_amount,running_balance,currency,
                clearing_document,gl_account,special_gl,offsetting_type
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                code, customer_master[code],
                str(r.get("Posting Date","")), str(r.get("Document Date","")),
                str(r.get("Document Type","")), str(r.get("Document Number","")),
                str(r.get("Text","")), float(r.get("Debit") or 0), float(r.get("Credit") or 0),
                float(r.get("Net Amount") or 0), float(r.get("Running Balance") or 0),
                str(r.get("Currency","")), str(r.get("Clearing Document","")),
                str(r.get("G/L Account","")), str(r.get("Special G/L","")),
                str(r.get("Offsetting Acct Type",""))
            ))
        for r in seed.get("duplicate_docs", []):
            code = clean_customer_code(r.get("Customer Code"))
            if code not in customer_master:
                continue
            cur.execute("""INSERT INTO duplicate_documents(
                customer_code,customer_name,posting_date,document_date,document_type,
                document_number,text,debit,credit,net_amount,currency,clearing_document,
                gl_account,special_gl,offsetting_type,duplicate_occurrences
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                code, customer_master[code],
                str(r.get("Posting Date","")), str(r.get("Document Date","")),
                str(r.get("Document Type","")), str(r.get("Document Number","")),
                str(r.get("Text","")), float(r.get("Debit") or 0), float(r.get("Credit") or 0),
                float(r.get("Net Amount") or 0), str(r.get("Currency","")),
                str(r.get("Clearing Document","")), str(r.get("G/L Account","")),
                str(r.get("Special G/L","")), str(r.get("Offsetting Acct Type","")),
                int(r.get("Duplicate Occurrences") or 0)
            ))
        conn.commit()

    had_cash_before_seed = cur.execute("SELECT COUNT(*) FROM cash_ledger").fetchone()[0] > 0
    seed_cash_data_from_excel(cur)
    seed_return_approvers_from_excel(cur)
    seed_cashier_closing_from_excel(cur)

    # One-time correction for databases created by older versions where Head Cash
    # Debit and Credit were stored on the opposite sides. Fresh databases are
    # already seeded with the corrected presentation and are only marked here.
    migration_key = "v23_head_cash_debit_credit_fix"
    already_fixed = cur.execute("SELECT 1 FROM app_migrations WHERE migration_key=?", (migration_key,)).fetchone()
    if not already_fixed:
        if had_cash_before_seed:
            cur.execute("UPDATE cash_ledger SET debit=credit, credit=debit WHERE cash_type='Head Cash'")
            cur.execute("UPDATE deleted_cash_entries SET debit=credit, credit=debit WHERE cash_type='Head Cash'")
        cur.execute("INSERT OR REPLACE INTO app_migrations(migration_key,applied_at) VALUES(?,?)",
                    (migration_key, datetime.now().isoformat(timespec="seconds")))
    conn.commit()

    # Recover previously created users from the permanent sidecar backup before
    # creating the default admin. This protects accounts even if a database file
    # was manually replaced.
    restore_users_from_backup(conn)
    conn.commit()

    if cur.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        cur.execute("""INSERT INTO users(full_name,username,password_hash,role_name,user_type,store_access,permissions,status,created_at)
                       VALUES(?,?,?,?,?,?,?,?,?)""",(
            "Rahat Ullah","Rahat",generate_password_hash("Rahat@0031"),"Super Admin","Local","ALL",
            json.dumps(PERMISSIONS),"Active",datetime.now().isoformat(timespec="seconds")
        ))
        conn.commit()
    backup_users(conn)
    conn.close()

def require_login():
    return session.get("logged_in") is True

@app.route("/")
def home():
    return send_from_directory(BASE, "index.html")

@app.before_request
def load_current_user():
    if request.path.startswith("/static/") or request.path in ("/", "/api/login"):
        return None
    uid=session.get("user_id")
    if not uid:
        return jsonify({"error":"Login required"}),401
    conn=auth_db(); row=conn.execute("SELECT * FROM users WHERE id=? AND status='Active'",(uid,)).fetchone(); conn.close()
    if not row:
        session.clear(); return jsonify({"error":"Account inactive"}),401
    g.user=dict(row)

def user_permissions():
    if not getattr(g,"user",None): return set()
    # Rahat/Admin is the permanent owner account and always retains full access.
    if g.user.get("role_name")=="Super Admin" or str(g.user.get("username") or "").strip().lower()=="rahat":
        return set(PERMISSIONS)
    try: return set(json.loads(g.user.get("permissions") or "[]"))
    except Exception: return set()

LEGACY_PERMISSION_MAP = {
    "dashboard_view":["dashboard_view"],
    "customer_view":["corporate_customers_view"], "customer_import":["corporate_customers_add","corporate_customers_edit"], "customer_delete":["corporate_customers_delete"],
    "duplicate_view":["documents_data_view"], "duplicate_delete":["documents_data_delete"],
    "vendor_view":["vendor_management_view"], "vendor_add":["vendor_management_add"], "vendor_edit":["vendor_management_edit"], "vendor_delete":["vendor_management_delete"],
    "ledger_view":["vendor_management_view","corporate_customers_view"], "ledger_add":["vendor_management_add","corporate_customers_add"], "ledger_delete":["vendor_management_delete","corporate_customers_delete"],
    "cash_view":["head_cash_view","petty_cash_view","lost_found_view","theft_view"], "cash_import":["head_cash_add","petty_cash_add","lost_found_add","theft_add"], "cash_delete":["head_cash_delete","petty_cash_delete","lost_found_delete","theft_delete"],
    "cashier_view":["cashier_closing_view"], "cashier_import":["cashier_closing_add","cashier_closing_edit"], "cashier_delete":["cashier_closing_delete"],
    "return_view":["return_counter_view"], "return_import":["return_counter_add","return_counter_edit"], "return_delete":["return_counter_delete"],
    "export_data":["documents_data_export","reports_export"], "backup_restore":["documents_data_edit"],
    "user_manage":["security_users_view","security_users_add","security_users_edit","security_users_delete"], "audit_view":["reports_view"]
}

def has_permission(name):
    perms=user_permissions()
    return name in perms or any(p in perms for p in LEGACY_PERMISSION_MAP.get(name,[]))

def require_permission(name):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if not has_permission(name): return jsonify({"error":"Permission denied"}),403
            return fn(*args,**kwargs)
        return wrapper
    return deco

def audit(action,module,details=""):
    try:
        conn=auth_db(); conn.execute("INSERT INTO audit_log(username,action,module,details,created_at,ip_address) VALUES(?,?,?,?,?,?)",(
            getattr(g,"user",{}).get("username",session.get("username","System")),action,module,details,
            datetime.now().isoformat(timespec="seconds"),request.remote_addr or "")); conn.commit(); conn.close()
    except Exception: pass

@app.post("/api/login")
def login():
    data=request.get_json(force=True)
    username=str(data.get("username") or "").strip()
    conn=auth_db(); row=conn.execute("SELECT * FROM users WHERE lower(username)=lower(?)",(username,)).fetchone()
    ok=bool(row and row["status"]=="Active" and check_password_hash(row["password_hash"],str(data.get("password") or "")))
    if ok:
        session.clear(); session["user_id"]=row["id"]; session["username"]=row["username"]
        conn.execute("UPDATE users SET last_login=? WHERE id=?",(datetime.now().isoformat(timespec="seconds"),row["id"])); conn.commit()
        user={"id":row["id"],"full_name":row["full_name"],"username":row["username"],"role_name":row["role_name"],"permissions":PERMISSIONS if row["role_name"]=="Super Admin" else json.loads(row["permissions"] or "[]")}
    else: user=None
    conn.close()
    return jsonify({"ok":ok,"user":user})

@app.get("/api/me")
def me():
    return jsonify({"id":g.user["id"],"full_name":g.user["full_name"],"username":g.user["username"],"role_name":g.user["role_name"],"user_type":g.user["user_type"],"store_access":g.user["store_access"],"permissions":sorted(user_permissions())})

@app.post("/api/logout")
def logout():
    audit("Logout","Authentication")
    session.clear()
    return jsonify({"ok":True})

@app.get("/api/stores")
def list_stores():
    selected = active_store_code()
    access = str(g.user.get("store_access") or "ALL").upper()
    allowed = STORES.items() if access == "ALL" else [(access, STORES.get(access, access))]
    return jsonify({"active_store": selected, "stores":[{"code":c,"name":n} for c,n in allowed]})

@app.post("/api/active-store")
def set_active_store():
    access = str(g.user.get("store_access") or "ALL").upper()
    code = str((request.get_json(silent=True) or {}).get("store_code") or "").upper()
    if code not in STORES:
        return jsonify({"error":"Invalid store"}),400
    if access != "ALL" and access != code:
        return jsonify({"error":"Store access denied"}),403
    session["active_store"] = code
    return jsonify({"ok":True,"store_code":code,"store_name":STORES[code]})

@app.get("/api/dashboard")
@require_permission("dashboard_view")
def dashboard():
    conn = db()
    cur = conn.cursor()
    customer_count = cur.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    total_debit = cur.execute("SELECT COALESCE(SUM(debit),0) FROM customer_ledger").fetchone()[0]
    total_credit = cur.execute("SELECT COALESCE(SUM(credit),0) FROM customer_ledger").fetchone()[0]
    duplicate_count = cur.execute("SELECT COUNT(DISTINCT document_number) FROM duplicate_documents").fetchone()[0]
    vendor_count = cur.execute("SELECT COUNT(*) FROM vendors").fetchone()[0]
    vendor_pending = cur.execute("SELECT COALESCE(SUM(debit-credit),0) FROM vendor_ledger").fetchone()[0]
    head_balance = cur.execute("SELECT COALESCE(SUM(credit-debit),0) FROM cash_ledger WHERE cash_type='Head Cash'").fetchone()[0]
    petty_balance = cur.execute("SELECT COALESCE(SUM(debit-credit),0) FROM cash_ledger WHERE cash_type='Petty Cash'").fetchone()[0]
    conn.close()
    return jsonify({
        "store_code": active_store_code(),
        "store_name": STORES.get(active_store_code(), active_store_code()),
        "customer_count": customer_count,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "customer_balance": total_debit-total_credit,
        "duplicate_documents": duplicate_count,
        "vendor_count": vendor_count,
        "vendor_pending": vendor_pending,
        "head_cash_balance": head_balance,
        "petty_cash_balance": petty_balance
    })



@app.get("/api/ledger-summary")
def ledger_summary():
    module = request.args.get("module", "").strip()
    q = request.args.get("q", "").strip()
    code = request.args.get("code", "").strip()
    date_from = request.args.get("from", "").strip()
    date_to = request.args.get("to", "").strip()
    conn = db()
    params = []
    if module == "corporate":
        sql = "SELECT COALESCE(SUM(debit),0) debit, COALESCE(SUM(credit),0) credit FROM customer_ledger WHERE 1=1"
        if q:
            sql += " AND (customer_code LIKE ? OR customer_name LIKE ? OR document_number LIKE ? OR text LIKE ?)"
            like=f"%{q}%"; params += [like,like,like,like]
        if code: sql += " AND customer_code=?"; params.append(code)
        if date_from: sql += " AND posting_date>=?"; params.append(date_from)
        if date_to: sql += " AND posting_date<=?"; params.append(date_to)
    elif module == "duplicates":
        sql = "SELECT COALESCE(SUM(debit),0) debit, COALESCE(SUM(credit),0) credit FROM duplicate_documents WHERE 1=1"
        if q:
            sql += " AND (customer_code LIKE ? OR customer_name LIKE ? OR document_number LIKE ?)"
            like=f"%{q}%"; params += [like,like,like]
    elif module == "vendor":
        sql = "SELECT COALESCE(SUM(debit),0) debit, COALESCE(SUM(credit),0) credit FROM vendor_ledger WHERE 1=1"
        if code: sql += " AND vendor_code=?"; params.append(code)
        if q:
            sql += " AND (vendor_code LIKE ? OR vendor_name LIKE ? OR document_number LIKE ? OR description LIKE ? OR gate_pass LIKE ? OR bilty_no LIKE ?)"
            like=f"%{q}%"; params += [like]*6
    elif module in ("Head Cash", "Petty Cash"):
        sql = "SELECT COALESCE(SUM(debit),0) debit, COALESCE(SUM(credit),0) credit FROM cash_ledger WHERE cash_type=?"
        params=[module]
        if q:
            sql += " AND (document_number LIKE ? OR description LIKE ?)"
            like=f"%{q}%"; params += [like,like]
        if date_from: sql += " AND posting_date>=?"; params.append(date_from)
        if date_to: sql += " AND posting_date<=?"; params.append(date_to)
    elif module == "Deleted Entries":
        sql = "SELECT COALESCE(SUM(debit),0) debit, COALESCE(SUM(credit),0) credit FROM deleted_cash_entries WHERE 1=1"
        if q:
            sql += " AND (cash_type LIKE ? OR document_number LIKE ? OR description LIKE ?)"
            like=f"%{q}%"; params += [like,like,like]
    else:
        conn.close(); return jsonify({"error":"Invalid module"}),400
    row=conn.execute(sql,params).fetchone(); conn.close()
    debit=float(row["debit"] or 0); credit=float(row["credit"] or 0)
    balance = (credit-debit) if module == "Head Cash" else (debit-credit)
    return jsonify({"debit":debit,"credit":credit,"balance":balance})

@app.get("/api/customers")
@require_permission("customer_view")
def customers():
    conn = db()
    rows = conn.execute("SELECT code,name FROM customers ORDER BY code").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.get("/api/customer-ledger")
@require_permission("customer_view")
def customer_ledger():
    q = request.args.get("q","").strip()
    code = request.args.get("code","").strip()
    date_from = request.args.get("from","").strip()
    date_to = request.args.get("to","").strip()
    sql = "SELECT * FROM customer_ledger WHERE 1=1"
    params = []
    if q:
        sql += " AND (customer_code LIKE ? OR customer_name LIKE ? OR document_number LIKE ? OR text LIKE ?)"
        like = f"%{q}%"; params += [like,like,like,like]
    if code:
        sql += " AND customer_code=?"; params.append(code)
    if date_from:
        sql += " AND posting_date>=?"; params.append(date_from)
    if date_to:
        sql += " AND posting_date<=?"; params.append(date_to)
    sql += " ORDER BY customer_code, posting_date, document_number LIMIT 5000"
    conn = db()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.post("/api/customer-ledger/bulk")
@require_permission("customer_import")
def bulk_customer_ledger():
    data = request.get_json(force=True) or {}
    entries = data.get("entries", [])
    if not isinstance(entries, list) or not entries:
        return jsonify({"error":"No entries supplied"}), 400

    conn = db(); cur = conn.cursor()
    master = {clean_customer_code(r["code"]): r["name"] for r in cur.execute("SELECT code,name FROM customers")}
    existing_docs = {str(r[0] or "").strip() for r in cur.execute(
        "SELECT document_number FROM customer_ledger WHERE TRIM(COALESCE(document_number,''))<>''"
    )}
    duplicate_docs = {str(r[0] or "").strip() for r in cur.execute(
        "SELECT document_number FROM duplicate_documents WHERE TRIM(COALESCE(document_number,''))<>''"
    )}
    seen_in_batch = set()
    inserted = 0; duplicate_rows = 0; skipped = []
    today = datetime.now().strftime("%Y-%m-%d")

    for i, row in enumerate(entries, start=1):
        code = clean_customer_code(row.get("code"))
        doc = str(row.get("document_number") or "").strip()
        entry_date = str(row.get("date") or today).strip()
        debit = n(row.get("debit")); credit = n(row.get("credit"))
        description = str(row.get("description") or "").strip()

        if not doc:
            skipped.append({"row":i,"code":code,"reason":"Document No. is required"})
            continue
        if code not in master:
            skipped.append({"row":i,"code":code,"reason":"Customer code not found in master"})
            continue
        if debit == 0 and credit == 0:
            skipped.append({"row":i,"code":code,"reason":"Debit or Credit amount is required"})
            continue

        values = (code, master[code], entry_date, entry_date, "MANUAL", doc,
                  description, debit, credit, debit-credit, "PKR", "", "", "", "MANUAL")
        existing_same = cur.execute(
            "SELECT id FROM customer_ledger WHERE customer_code=? AND document_number=? ORDER BY id LIMIT 1",
            (code, doc)
        ).fetchone()
        if existing_same and doc not in seen_in_batch:
            cur.execute("""UPDATE customer_ledger SET customer_name=?,posting_date=?,document_date=?,document_type=?,
                text=?,debit=?,credit=?,net_amount=?,currency=?,clearing_document=?,gl_account=?,special_gl=?,offsetting_type=?
                WHERE id=?""", (
                master[code], entry_date, entry_date, "MANUAL", description, debit, credit, debit-credit,
                "PKR", "", "", "", "MANUAL", existing_same["id"]
            ))
            inserted += 1
        else:
            is_duplicate = doc in existing_docs or doc in duplicate_docs or doc in seen_in_batch
            if is_duplicate:
                occurrences = cur.execute(
                    "SELECT COUNT(*) FROM customer_ledger WHERE document_number=?", (doc,)
                ).fetchone()[0] + cur.execute(
                    "SELECT COUNT(*) FROM duplicate_documents WHERE document_number=?", (doc,)
                ).fetchone()[0] + 1
                cur.execute("""INSERT INTO duplicate_documents(
                    customer_code,customer_name,posting_date,document_date,document_type,
                    document_number,text,debit,credit,net_amount,currency,clearing_document,
                    gl_account,special_gl,offsetting_type,duplicate_occurrences
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", values + (occurrences,))
                duplicate_rows += 1
                duplicate_docs.add(doc)
            else:
                cur.execute("""INSERT INTO customer_ledger(
                    customer_code,customer_name,posting_date,document_date,document_type,
                    document_number,text,debit,credit,net_amount,running_balance,currency,
                    clearing_document,gl_account,special_gl,offsetting_type
                ) VALUES(?,?,?,?,?,?,?,?,?,?,0,?,?,?,?,?)""", values)
                inserted += 1
                existing_docs.add(doc)
        seen_in_batch.add(doc)

    # Store a correct customer-wise running balance in chronological/id order.
    affected = {clean_customer_code(r.get("code")) for r in entries if clean_customer_code(r.get("code")) in master}
    for code in affected:
        balance = 0.0
        rows = cur.execute(
            "SELECT id,debit,credit FROM customer_ledger WHERE customer_code=? ORDER BY posting_date, id", (code,)
        ).fetchall()
        for ledger_row in rows:
            balance += n(ledger_row["debit"]) - n(ledger_row["credit"])
            cur.execute("UPDATE customer_ledger SET running_balance=? WHERE id=?", (balance, ledger_row["id"]))

    conn.commit(); create_automatic_backup(force=True); conn.close()
    audit("Bulk Add","Corporate Ledger",f"Inserted {inserted}; duplicates {duplicate_rows}; skipped {len(skipped)}")
    return jsonify({"ok":True,"inserted":inserted,"duplicate_rows":duplicate_rows,"skipped":skipped})

@app.get("/api/duplicates")
@require_permission("duplicate_view")
def duplicates():
    q = request.args.get("q","").strip()
    sql = "SELECT * FROM duplicate_documents"
    params = []
    if q:
        sql += " WHERE customer_code LIKE ? OR customer_name LIKE ? OR document_number LIKE ?"
        like = f"%{q}%"; params = [like,like,like]
    sql += " ORDER BY document_number, customer_code LIMIT 5000"
    conn = db()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

def norm(s):
    return "".join(ch.lower() for ch in str(s or "") if ch.isalnum())

def excel_rows(ws):
    values = list(ws.iter_rows(values_only=True))
    if not values: return []
    best_i, best_score = 0, -1
    words = ["document","customer","posting","offsetting","amount","debit","credit"]
    for i,line in enumerate(values[:20]):
        score = sum(1 for c in line if any(w in norm(c) for w in words))
        if score > best_score:
            best_i,best_score=i,score
    headers = [str(x or "").strip() for x in values[best_i]]
    out=[]
    for line in values[best_i+1:]:
        if not any(v not in (None,"") for v in line): continue
        out.append({headers[i] if i<len(headers) else f"Column{i+1}": line[i] for i in range(len(line))})
    return out

def pick(row,*aliases):
    nrow={norm(k):v for k,v in row.items()}
    for a in aliases:
        if norm(a) in nrow: return nrow[norm(a)]
    return ""

def n(v):
    try: return float(str(v).replace(",",""))
    except: return 0.0

def d(v):
    if hasattr(v,"strftime"): return v.strftime("%Y-%m-%d")
    return str(v or "")

@app.post("/api/import-corporate")
@require_permission("customer_import")
def import_corporate():
    file = request.files.get("file")
    if not file:
        return jsonify({"error":"No file selected"}),400
    path=UPLOADS/file.filename
    file.save(path)
    wb=load_workbook(path,data_only=True,read_only=False)
    ws=wb[wb.sheetnames[0]]
    rows=excel_rows(ws)

    conn=db(); cur=conn.cursor()
    customer_names={clean_customer_code(r["code"]):r["name"] for r in cur.execute("SELECT code,name FROM customers").fetchall()}
    valid_codes=set(customer_names)

    normalized=[]
    skipped_non_customer=0
    for r in rows:
        doc=str(pick(r,"Document Number","Document No","Doc Number") or "").strip()
        doc_type=str(pick(r,"Document Type","Doc Type") or "").strip()
        if not doc or doc=="0" or not doc_type or doc_type=="0":
            continue
        code=find_master_customer_code(r, valid_codes)
        if not code:
            skipped_non_customer += 1
            continue
        amount=n(pick(r,"Amount in Local Currency","Net Amount","Amount"))
        normalized.append({
            "code":code,
            "name":customer_names[code],
            "posting_date":d(pick(r,"Posting Date")),
            "document_date":d(pick(r,"Document Date")),
            "document_type":doc_type,
            "document_number":doc,
            "text":str(pick(r,"Text","Description") or ""),
            "debit":amount if amount>0 else n(pick(r,"Debit Amount","Debit")),
            "credit":-amount if amount<0 else n(pick(r,"Credit Amount","Credit")),
            "net":amount,
            "currency":str(pick(r,"Local Currency","Currency") or "PKR"),
            "clearing_document":str(pick(r,"Clearing Document") or ""),
            "gl_account":str(pick(r,"G/L Account","GL Account") or ""),
            "special_gl":str(pick(r,"Special G/L Ind.","Special GL") or ""),
            "offset_type":str(pick(r,"Offsetting Acct Type") or "")
        })
    from collections import Counter
    counts=Counter(x["document_number"] for x in normalized)
    duplicates={doc for doc,c in counts.items() if c>1}
    imported=updated=dups=0
    affected_codes=set()
    for x in normalized:
        name=customer_names[x["code"]]
        affected_codes.add(x["code"])
        existing = cur.execute(
            "SELECT id FROM customer_ledger WHERE customer_code=? AND document_number=? ORDER BY id LIMIT 1",
            (x["code"], x["document_number"])
        ).fetchone()
        if existing and x["document_number"] not in duplicates:
            cur.execute("""UPDATE customer_ledger SET customer_name=?,posting_date=?,document_date=?,document_type=?,
                text=?,debit=?,credit=?,net_amount=?,currency=?,clearing_document=?,gl_account=?,special_gl=?,offsetting_type=?
                WHERE id=?""",(
                name,x["posting_date"],x["document_date"],x["document_type"],x["text"],x["debit"],x["credit"],
                x["net"],x["currency"],x["clearing_document"],x["gl_account"],x["special_gl"],x["offset_type"],existing["id"]
            )); updated+=1
        elif x["document_number"] in duplicates:
            cur.execute("""INSERT INTO duplicate_documents(
                customer_code,customer_name,posting_date,document_date,document_type,
                document_number,text,debit,credit,net_amount,currency,clearing_document,
                gl_account,special_gl,offsetting_type,duplicate_occurrences
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
                x["code"],name,x["posting_date"],x["document_date"],x["document_type"],
                x["document_number"],x["text"],x["debit"],x["credit"],x["net"],x["currency"],
                x["clearing_document"],x["gl_account"],x["special_gl"],x["offset_type"],counts[x["document_number"]]
            )); dups+=1
        else:
            cur.execute("""INSERT INTO customer_ledger(
                customer_code,customer_name,posting_date,document_date,document_type,
                document_number,text,debit,credit,net_amount,running_balance,currency,
                clearing_document,gl_account,special_gl,offsetting_type
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
                x["code"],name,x["posting_date"],x["document_date"],x["document_type"],
                x["document_number"],x["text"],x["debit"],x["credit"],x["net"],0,x["currency"],
                x["clearing_document"],x["gl_account"],x["special_gl"],x["offset_type"]
            )); imported+=1
    for code in affected_codes:
        balance=0.0
        for row in cur.execute("SELECT id,debit,credit FROM customer_ledger WHERE customer_code=? ORDER BY posting_date,id",(code,)).fetchall():
            balance += n(row["debit"]) - n(row["credit"])
            cur.execute("UPDATE customer_ledger SET running_balance=? WHERE id=?",(balance,row["id"]))
    conn.commit(); create_automatic_backup(force=True); conn.close()
    return jsonify({
        "ok":True,
        "imported":imported,
        "updated":updated,
        "duplicate_rows":dups,
        "duplicate_documents":len(duplicates),
        "skipped_non_customer_rows":skipped_non_customer
    })

@app.get("/api/cash-ledger")
@require_permission("cash_view")
def cash_ledger_api():
    cash_type=request.args.get("type","Head Cash")
    q=str(request.args.get("q") or "").strip()
    date_from=str(request.args.get("from") or "").strip(); date_to=str(request.args.get("to") or "").strip()
    conn=db(); sql="SELECT * FROM cash_ledger WHERE cash_type=?"; params=[cash_type]
    if q:
        sql+=" AND (document_number LIKE ? OR description LIKE ?)"; params += [f"%{q}%",f"%{q}%"]
    if date_from: sql+=" AND posting_date>=?"; params.append(date_from)
    if date_to: sql+=" AND posting_date<=?"; params.append(date_to)
    sql+=" ORDER BY posting_date,document_number,id LIMIT 10000"
    rows=[dict(r) for r in conn.execute(sql,params).fetchall()]; conn.close(); return jsonify(rows)

@app.get("/api/special-cash")
@require_permission("cash_view")
def special_cash_api():
    category=(request.args.get("category") or "").strip().lower()
    q=(request.args.get("q") or "").strip()
    date_from=(request.args.get("from") or "").strip(); date_to=(request.args.get("to") or "").strip()
    if category not in ("lost_found","theft"):
        return jsonify({"error":"Invalid category"}),400
    conn=db(); sql="SELECT * FROM cash_ledger WHERE cash_type='Head Cash'"; params=[]
    # Read directly from the Head Cash ledger so old records and every new
    # matching entry appear automatically.  Historical descriptions use many
    # spellings, for example Lost/Found, Lost & Found, Lost and Found and
    # Theft Recovery, therefore matching is deliberately flexible.
    search_text = "LOWER(COALESCE(description,'') || ' ' || COALESCE(document_number,''))"
    if category=="lost_found":
        sql += f" AND (({search_text} LIKE ? AND {search_text} LIKE ?) OR {search_text} LIKE ? OR {search_text} LIKE ?)"
        params += ["%lost%", "%found%", "%lost/found%", "%lost&found%"]
    else:
        sql += f" AND ({search_text} LIKE ? OR {search_text} LIKE ? OR {search_text} LIKE ?)"
        params += ["%theft%", "%stolen%", "%cash recovery%"]
    if q:
        sql += " AND (document_number LIKE ? OR description LIKE ?)"; params += [f"%{q}%",f"%{q}%"]
    if date_from: sql += " AND posting_date>=?"; params.append(date_from)
    if date_to: sql += " AND posting_date<=?"; params.append(date_to)
    sql += " ORDER BY posting_date,document_number,id LIMIT 10000"
    rows=[dict(r) for r in conn.execute(sql,params).fetchall()]; conn.close()
    return jsonify(rows)

@app.get("/api/deleted-cash")
@require_permission("cash_view")
def deleted_cash_api():
    q=str(request.args.get("q") or "").strip(); conn=db(); sql="SELECT * FROM deleted_cash_entries WHERE 1=1"; params=[]
    if q: sql+=" AND (document_number LIKE ? OR description LIKE ? OR cash_type LIKE ?)"; params=[f"%{q}%"]*3
    sql+=" ORDER BY id DESC LIMIT 10000"; rows=[dict(r) for r in conn.execute(sql,params).fetchall()]; conn.close(); return jsonify(rows)

@app.post("/api/cash-ledger")
@require_permission("cash_import")
def add_cash_entry():
    d=request.get_json(force=True); cash_type=d.get("cash_type")
    if cash_type not in ("Head Cash","Petty Cash"): return jsonify({"error":"Invalid cash type"}),400
    conn=db(); conn.execute("INSERT INTO cash_ledger(cash_type,document_number,document_date,posting_date,description,debit,credit,created_at) VALUES(?,?,?,?,?,?,?,?)",(cash_type,str(d.get("document_number") or ""),str(d.get("document_date") or ""),str(d.get("posting_date") or ""),str(d.get("description") or ""),float(d.get("debit") or 0),float(d.get("credit") or 0),datetime.now().isoformat(timespec="seconds"))); conn.commit(); conn.close(); audit("Add",cash_type,"Manual cash entry"); return jsonify({"ok":True})

@app.post("/api/cash-ledger/import")
@require_permission("cash_import")
def import_cash_file():
    file=request.files.get("file"); cash_type=request.form.get("cash_type")
    if not file or cash_type not in ("Head Cash","Petty Cash"): return jsonify({"error":"File and cash type required"}),400
    wb=load_workbook(file,read_only=True,data_only=True); ws=wb.active; rows=ws.iter_rows(values_only=True); first=next(rows,None); hdr=[str(x or "").strip().lower() for x in (first or [])]
    has_header=any("document" in x or "posting" in x for x in hdr)
    data_rows=rows if has_header else iter([first]+list(rows))
    def idx(names,default):
        for n in names:
            for i,h in enumerate(hdr):
                if n in h:return i
        return default
    di=idx(["document number","document no"],0); ddi=idx(["document date"],1); pdi=idx(["posting date"],2); desci=idx(["description","text"],3); debi=idx(["debit"],4); credi=idx(["credit"],5)
    conn=db(); batch=[]; inserted=0
    for r in data_rows:
        if not r: continue
        vals=list(r)+[None]*8
        def ds(v): return v.strftime("%Y-%m-%d") if hasattr(v,"strftime") else str(v or "")
        try: debit=float(vals[debi] or 0)
        except: debit=0
        try: credit=float(vals[credi] or 0)
        except: credit=0
        if not any([vals[di],vals[desci],debit,credit]): continue
        batch.append((cash_type,str(vals[di] or ""),ds(vals[ddi]),ds(vals[pdi]),str(vals[desci] or ""),debit,credit,datetime.now().isoformat(timespec="seconds")))
        if len(batch)>=2000:
            conn.executemany("INSERT INTO cash_ledger(cash_type,document_number,document_date,posting_date,description,debit,credit,created_at) VALUES(?,?,?,?,?,?,?,?)",batch); inserted+=len(batch); batch=[]
    if batch: conn.executemany("INSERT INTO cash_ledger(cash_type,document_number,document_date,posting_date,description,debit,credit,created_at) VALUES(?,?,?,?,?,?,?,?)",batch); inserted+=len(batch)
    conn.commit(); conn.close(); wb.close(); audit("Import",cash_type,f"{inserted} rows"); return jsonify({"inserted":inserted})

@app.post("/api/cash-ledger/delete-selected")
@require_permission("cash_delete")
def delete_cash_selected():
    d=request.get_json(force=True); ids=[int(x) for x in d.get("ids",[]) if str(x).isdigit()]
    if not ids:return jsonify({"error":"No rows selected"}),400
    conn=db(); q=','.join('?' for _ in ids); rows=conn.execute(f"SELECT * FROM cash_ledger WHERE id IN ({q})",ids).fetchall(); now=datetime.now().isoformat(timespec="seconds"); user=g.user.get("username","")
    for r in rows: conn.execute("INSERT INTO deleted_cash_entries(original_id,cash_type,document_number,document_date,posting_date,description,debit,credit,deleted_by,deleted_at,delete_reason) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(r['id'],r['cash_type'],r['document_number'],r['document_date'],r['posting_date'],r['description'],r['debit'],r['credit'],user,now,"Deleted by user"))
    conn.execute(f"DELETE FROM cash_ledger WHERE id IN ({q})",ids); conn.commit(); conn.close(); audit("Delete","Cash Ledger",f"Moved {len(rows)} rows to Deleted Entries"); return jsonify({"deleted":len(rows)})

@app.post("/api/deleted-cash/restore/<int:row_id>")
@require_permission("cash_delete")
def restore_cash(row_id):
    conn=db(); r=conn.execute("SELECT * FROM deleted_cash_entries WHERE id=?",(row_id,)).fetchone()
    if not r: conn.close(); return jsonify({"error":"Record not found"}),404
    conn.execute("INSERT INTO cash_ledger(cash_type,document_number,document_date,posting_date,description,debit,credit,created_at) VALUES(?,?,?,?,?,?,?,?)",(r['cash_type'],r['document_number'],r['document_date'],r['posting_date'],r['description'],r['debit'],r['credit'],datetime.now().isoformat(timespec="seconds"))); conn.execute("DELETE FROM deleted_cash_entries WHERE id=?",(row_id,)); conn.commit(); conn.close(); audit("Restore","Cash Ledger",str(r['document_number'])); return jsonify({"ok":True})

@app.get("/api/export-cash/<cash_type>")
@require_permission("export_data")
def export_cash(cash_type):
    if cash_type not in ("Head Cash","Petty Cash","Deleted Entries"): return jsonify({"error":"Invalid type"}),400
    conn=db()
    if cash_type=="Deleted Entries": rows=conn.execute("SELECT * FROM deleted_cash_entries ORDER BY cash_type,posting_date,document_number").fetchall()
    else: rows=conn.execute("SELECT * FROM cash_ledger WHERE cash_type=? ORDER BY posting_date,document_number,id",(cash_type,)).fetchall()
    conn.close(); wb=Workbook(); ws=wb.active; ws.title=cash_type[:31]
    ws.append(["Document No.","Document Date","Posting Date","Description","Debit Amount","Credit Amount","Balance"])
    bal=0
    for r in rows:
        bal += (float(r['credit'] or 0)-float(r['debit'] or 0)) if cash_type=='Head Cash' else (float(r['debit'] or 0)-float(r['credit'] or 0)); ws.append([r['document_number'],r['document_date'],r['posting_date'],r['description'],r['debit'],r['credit'],bal])
    for c in ws[1]: c.font=c.font.copy(bold=True)
    bio=io.BytesIO(); wb.save(bio); bio.seek(0); return send_file(bio,as_attachment=True,download_name=cash_type.replace(' ','_')+'_Ledger.xlsx')



def return_date_clause(args):
    period=(args.get('period') or 'all').strip().lower(); params=[]; clauses=[]
    if period=='day' and args.get('date'):
        clauses.append('entry_date=?'); params.append(args.get('date'))
    elif period=='week' and args.get('week'):
        try:
            y,w=map(int,args.get('week').split('-W')); start=datetime.fromisocalendar(y,w,1).date(); end=start+timedelta(days=6)
            clauses += ['entry_date>=?','entry_date<=?']; params += [start.isoformat(),end.isoformat()]
        except: pass
    elif period=='month' and args.get('month'):
        clauses.append('substr(entry_date,1,7)=?'); params.append(args.get('month'))
    elif period=='year' and args.get('year'):
        clauses.append('substr(entry_date,1,4)=?'); params.append(args.get('year'))
    if args.get('from'):
        clauses.append('entry_date>=?'); params.append(args.get('from'))
    if args.get('to'):
        clauses.append('entry_date<=?'); params.append(args.get('to'))
    return clauses,params

@app.get('/api/return-approvers')
@require_permission('return_view')
def return_approvers_api():
    conn=db(); rows=[dict(r) for r in conn.execute("SELECT * FROM return_approvers WHERE active=1 ORDER BY management_name").fetchall()]; conn.close(); return jsonify(rows)

@app.get('/api/return-entries')
@require_permission('return_view')
def return_entries_api():
    clauses,params=return_date_clause(request.args); q=(request.args.get('q') or '').strip(); approval=(request.args.get('approval') or '').strip(); source=(request.args.get('source') or '').strip()
    sql='SELECT * FROM return_entries WHERE 1=1'
    for c in clauses: sql+=' AND '+c
    if q:
        sql+=" AND (customer_name LIKE ? OR trx_no LIKE ? OR return_trx LIKE ? OR item_description LIKE ? OR reason LIKE ?)"; params += [f'%{q}%']*5
    if approval: sql+=' AND approval_name=?'; params.append(approval)
    if source: sql+=' AND source_file=?'; params.append(source)
    sql+=' ORDER BY entry_date DESC,id DESC LIMIT 50000'
    conn=db(); rows=[dict(r) for r in conn.execute(sql,params).fetchall()]
    totals=dict(conn.execute('''SELECT COUNT(*) records,
        COALESCE(SUM(total_amount),0) total_amount,
        COALESCE(SUM(CASE WHEN lower(COALESCE(cash_voucher,'')) LIKE '%voucher%' THEN total_amount ELSE 0 END),0) voucher_amount,
        COALESCE(SUM(CASE WHEN lower(COALESCE(cash_voucher,'')) NOT LIKE '%voucher%' THEN total_amount ELSE 0 END),0) cash_amount,
        COALESCE(SUM(CASE WHEN system_status="Valid Approval" THEN total_amount ELSE 0 END),0) valid_amount,
        COALESCE(SUM(CASE WHEN system_status="Limit Exceeded" THEN total_amount ELSE 0 END),0) exceeded_amount,
        COALESCE(SUM(CASE WHEN system_status="Valid Approval" THEN 1 ELSE 0 END),0) valid_entries,
        COALESCE(SUM(CASE WHEN system_status="Limit Exceeded" THEN 1 ELSE 0 END),0) exceeded_entries
        FROM ('''+sql.replace(' LIMIT 50000','')+')',params).fetchone())
    sources=[r[0] for r in conn.execute("SELECT DISTINCT source_file FROM return_entries WHERE COALESCE(source_file,'')<>'' ORDER BY source_file").fetchall()]
    conn.close(); return jsonify({'rows':rows,'totals':totals,'sources':sources})

def return_status(conn,name,amount):
    r=conn.execute('SELECT designation,approval_limit,unlimited FROM return_approvers WHERE management_name=? AND active=1',(name,)).fetchone()
    if not r: return '',0,'Approval Not Found'
    limit=float(r['approval_limit'] or 0); status='Valid Approval' if r['unlimited'] or amount<=limit else 'Limit Exceeded'
    return r['designation'],limit,status

@app.post('/api/return-entries/bulk')
@require_permission('return_import')
def return_entries_bulk():
    entries=(request.get_json(force=True) or {}).get('entries') or []
    if not isinstance(entries,list) or not entries: return jsonify({'error':'No entries supplied'}),400
    conn=db(); now=datetime.now().isoformat(timespec='seconds'); inserted=0; invalid=[]
    sql="INSERT INTO return_entries(entry_date,serial_no,customer_name,contact_no,return_trx,trx_no,item_description,item_exp_date,qty,total_amount,reason,approval_name,designation,approval_limit,system_status,cash_voucher,received_by,cctv_time,trx_time,source_file,source_sheet,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
    for row_no,e in enumerate(entries,start=1):
        errors=[]
        if not isinstance(e,dict): invalid.append({'row':row_no,'errors':['Invalid row format']}); continue
        if not str(e.get('customer_name') or '').strip(): errors.append('Customer Name required')
        if not str(e.get('trx_no') or '').strip(): errors.append('Trx No required')
        if not str(e.get('item_description') or '').strip(): errors.append('Item Description required')
        try: amount=float(str(e.get('total_amount') or '0').replace(',',''))
        except: amount=0; errors.append('Total Amount must be numeric')
        try: qty=float(str(e.get('qty') or '0').replace(',',''))
        except: qty=0; errors.append('Qty must be numeric')
        approval=str(e.get('approval_name') or '').strip()
        designation,limit,status=return_status(conn,approval,amount)
        if approval and status=='Approval Not Found': errors.append('Approval Name not found')
        if errors:
            invalid.append({'row':row_no,'errors':errors,'data':e}); continue
        vals=(str(e.get('entry_date') or ''),str(e.get('serial_no') or ''),str(e.get('customer_name') or ''),str(e.get('contact_no') or ''),str(e.get('return_trx') or ''),str(e.get('trx_no') or ''),str(e.get('item_description') or ''),str(e.get('item_exp_date') or ''),qty,amount,str(e.get('reason') or ''),approval,designation,limit,status,str(e.get('cash_voucher') or ''),str(e.get('received_by') or ''),str(e.get('cctv_time') or ''),str(e.get('trx_time') or ''),'Excel Paste','Bulk Paste',now)
        conn.execute(sql,vals); inserted+=1
    conn.commit(); conn.close(); audit('Bulk Add','Return Counter',f'{inserted} valid entries; {len(invalid)} invalid')
    return jsonify({'inserted':inserted,'invalid_count':len(invalid),'invalid':invalid})

@app.post('/api/return-entries/import')
@require_permission('return_import')
def return_entries_import():
    file=request.files.get('file'); month=(request.form.get('month') or '').strip()
    if not file or not month: return jsonify({'error':'Excel file and month are required'}),400
    filename=Path(file.filename or 'Return_Report.xlsx').name; wb=load_workbook(file,read_only=True,data_only=True); conn=db(); now=datetime.now().isoformat(timespec='seconds'); inserted=0
    # Refresh management names/limits from every uploaded file.
    if 'Management' in wb.sheetnames:
        for row in wb['Management'].iter_rows(min_row=3,max_col=3,values_only=True):
            designation,name,limit=row
            if not name: continue
            unlimited=1 if str(limit or '').strip().lower()=='unlimited' or 'store manager' in str(designation or '').lower() else 0
            try: lim=float(limit or 0) if not unlimited else 0
            except: lim=0
            conn.execute('INSERT INTO return_approvers(designation,management_name,approval_limit,unlimited,active) VALUES(?,?,?,?,1) ON CONFLICT(management_name) DO UPDATE SET designation=excluded.designation,approval_limit=excluded.approval_limit,unlimited=excluded.unlimited,active=1',(str(designation or ''),str(name).strip(),lim,unlimited))
    sql="INSERT INTO return_entries(entry_date,serial_no,customer_name,contact_no,return_trx,trx_no,item_description,item_exp_date,qty,total_amount,reason,approval_name,designation,approval_limit,system_status,cash_voucher,received_by,cctv_time,trx_time,source_file,source_sheet,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
    for sheet in wb.sheetnames:
        if not sheet.isdigit(): continue
        day=int(sheet)
        try: entry_date=datetime.strptime(f'{month}-{day:02d}','%Y-%m-%d').date().isoformat()
        except: continue
        ws=wb[sheet]
        for row in ws.iter_rows(min_row=2,max_col=17,values_only=True):
            vals=list(row)+[None]*17
            if not any(v not in (None,'') for v in vals[:17]): continue
            try: amount=float(vals[8] or 0)
            except: amount=0
            try: qty=float(vals[7] or 0)
            except: qty=0
            approval=str(vals[10] or '').strip(); designation,limit,status=return_status(conn,approval,amount)
            def ds(v): return v.strftime('%Y-%m-%d') if hasattr(v,'strftime') else str(v or '')
            conn.execute(sql,(entry_date,str(vals[0] or ''),str(vals[1] or ''),str(vals[2] or ''),str(vals[3] or ''),str(vals[4] or ''),str(vals[5] or ''),ds(vals[6]),qty,amount,str(vals[9] or ''),approval,designation or str(vals[11] or ''),limit,status,str(vals[13] or ''),str(vals[14] or ''),str(vals[15] or ''),str(vals[16] or ''),filename,sheet,now)); inserted+=1
    conn.commit(); conn.close(); wb.close(); audit('Import','Return Counter',f'{filename}: {inserted} rows'); return jsonify({'inserted':inserted,'file':filename})

@app.post('/api/return-entries/delete-selected')
@require_permission('return_delete')
def return_delete_selected():
    ids=[int(x) for x in (request.get_json(force=True) or {}).get('ids',[]) if str(x).isdigit()]
    if not ids: return jsonify({'error':'No rows selected'}),400
    conn=db(); q=','.join('?'*len(ids)); conn.execute(f'DELETE FROM return_entries WHERE id IN ({q})',ids); conn.commit(); conn.close(); audit('Delete','Return Counter',f'{len(ids)} rows'); return jsonify({'deleted':len(ids)})

def _return_export_query(args):
    clauses,params=return_date_clause(args)
    q=(args.get('q') or '').strip(); approval=(args.get('approval') or '').strip(); source=(args.get('source') or '').strip()
    sql='SELECT * FROM return_entries WHERE 1=1'
    for c in clauses: sql+=' AND '+c
    if q:
        sql+=" AND (customer_name LIKE ? OR trx_no LIKE ? OR return_trx LIKE ? OR item_description LIKE ? OR reason LIKE ?)"; params += [f'%{q}%']*5
    if approval: sql+=' AND approval_name=?'; params.append(approval)
    if source: sql+=' AND source_file=?'; params.append(source)
    return sql+' ORDER BY entry_date,id',params

def _return_summary(rows):
    total=sum(float(r['total_amount'] or 0) for r in rows)
    voucher=sum(float(r['total_amount'] or 0) for r in rows if 'voucher' in str(r['cash_voucher'] or '').lower())
    valid=sum(float(r['total_amount'] or 0) for r in rows if r['system_status']=='Valid Approval')
    exceeded=sum(float(r['total_amount'] or 0) for r in rows if r['system_status']=='Limit Exceeded')
    return {'cash':total-voucher,'voucher':voucher,'total':total,'valid':valid,'exceeded':exceeded,
            'valid_count':sum(1 for r in rows if r['system_status']=='Valid Approval'),
            'exceeded_count':sum(1 for r in rows if r['system_status']=='Limit Exceeded'),'count':len(rows)}

@app.get('/api/export-return-entries')
@require_permission('return_counter_export')
def export_return_entries():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    sql,params=_return_export_query(request.args); conn=db(); rows=conn.execute(sql,params).fetchall(); conn.close(); totals=_return_summary(rows)
    out=Workbook(); ws=out.active; ws.title='Return Counter Report'; ws.sheet_view.showGridLines=False
    headers=['Date','S.No.','Customer Name','Contact No','Return Trx','Trx No','Item Description','Item Exp Date','Qty','Total Amount','Reason','Approval Name','Designation','Approval Limit','System Status','Cash/Voucher','Rcv by','CCTV/Time','Trx Time','Source File','Source Sheet']
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers)); ws['A1']='RETURN / EXCHANGE COUNTER DASHBOARD REPORT'
    ws['A1'].font=Font(size=18,bold=True,color='FFFFFF'); ws['A1'].fill=PatternFill('solid',fgColor='174E75'); ws['A1'].alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=30
    labels=[('Cash Amount',totals['cash']),('Voucher Amount',totals['voucher']),('Total Amount',totals['total']),('Valid Approval Amount',totals['valid']),('Exceeded Approval Amount',totals['exceeded']),('Valid Entries',totals['valid_count']),('Exceeded Entries',totals['exceeded_count']),('Total Entries',totals['count'])]
    for i,(label,value) in enumerate(labels):
        col=1+(i%4)*5; row=3+(i//4)*3
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+3); ws.cell(row,col,label)
        ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+3); ws.cell(row+1,col,value)
        ws.cell(row,col).font=Font(bold=True,color='5C7080'); ws.cell(row,col).alignment=Alignment(horizontal='center')
        ws.cell(row+1,col).font=Font(bold=True,size=14,color='173A54'); ws.cell(row+1,col).alignment=Alignment(horizontal='center')
        if 'Entries' not in label: ws.cell(row+1,col).number_format='Rs. #,##0.00'
        fill='EAF3F8' if 'Exceeded' not in label else 'FDECEC'
        for rr in (row,row+1):
            for cc in range(col,col+4): ws.cell(rr,cc).fill=PatternFill('solid',fgColor=fill)
    header_row=9; ws.append([])
    for c,h in enumerate(headers,1): ws.cell(header_row,c,h)
    for r in rows: ws.append([r['entry_date'],r['serial_no'],r['customer_name'],r['contact_no'],r['return_trx'],r['trx_no'],r['item_description'],r['item_exp_date'],r['qty'],r['total_amount'],r['reason'],r['approval_name'],r['designation'],r['approval_limit'],r['system_status'],r['cash_voucher'],r['received_by'],r['cctv_time'],r['trx_time'],r['source_file'],r['source_sheet']])
    thin=Side(style='thin',color='D8E1E8')
    for cell in ws[header_row]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='2F6F9F'); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=Border(bottom=thin)
    for row in ws.iter_rows(min_row=header_row+1,max_row=ws.max_row):
        for cell in row: cell.border=Border(bottom=thin); cell.alignment=Alignment(vertical='top',wrap_text=True)
        status=str(row[14].value or '')
        row[14].fill=PatternFill('solid',fgColor='DDF2E3' if status=='Valid Approval' else 'F9DADA'); row[14].font=Font(bold=True,color='176B2C' if status=='Valid Approval' else '9B111E')
        row[9].number_format='Rs. #,##0.00'; row[13].number_format='Rs. #,##0.00'
    widths=[12,9,24,16,15,15,32,14,9,16,30,22,20,16,18,15,18,18,14,22,14]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes=f'A{header_row+1}'; ws.auto_filter.ref=f'A{header_row}:U{ws.max_row}'; ws.print_title_rows=f'1:{header_row}'; ws.page_setup.orientation='landscape'; ws.page_setup.fitToWidth=1; ws.sheet_properties.pageSetUpPr.fitToPage=True
    bio=io.BytesIO(); out.save(bio); bio.seek(0); return send_file(bio,as_attachment=True,download_name=f'Return_Counter_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.get('/api/export-return-entries.pdf')
@require_permission('return_counter_print')
def export_return_entries_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    sql,params=_return_export_query(request.args); conn=db(); rows=conn.execute(sql,params).fetchall(); conn.close(); totals=_return_summary(rows)
    bio=io.BytesIO(); doc=SimpleDocTemplate(bio,pagesize=landscape(A3),rightMargin=22,leftMargin=22,topMargin=24,bottomMargin=22)
    styles=getSampleStyleSheet(); title=ParagraphStyle('title',parent=styles['Title'],fontSize=20,textColor=colors.HexColor('#174E75'),alignment=TA_CENTER,spaceAfter=12)
    small=ParagraphStyle('small',parent=styles['BodyText'],fontSize=6.5,leading=8)
    story=[Paragraph('RETURN / EXCHANGE COUNTER DASHBOARD REPORT',title)]
    cards=[['Cash Amount',f"Rs. {totals['cash']:,.2f}",'Voucher Amount',f"Rs. {totals['voucher']:,.2f}",'Total Amount',f"Rs. {totals['total']:,.2f}"],['Valid Approval',f"Rs. {totals['valid']:,.2f}",'Exceeded Approval',f"Rs. {totals['exceeded']:,.2f}",'Entries',f"{totals['count']:,} (Valid {totals['valid_count']:,} / Exceeded {totals['exceeded_count']:,})"]]
    ct=Table(cards,colWidths=[100,145,110,145,105,190],rowHeights=28)
    ct.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#EAF3F8')),('GRID',(0,0),(-1,-1),.5,colors.HexColor('#B9CBD8')),('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,-1),'RIGHT'),('ALIGN',(3,0),(3,-1),'RIGHT'),('ALIGN',(5,0),(5,-1),'RIGHT')]))
    story += [ct,Spacer(1,14)]
    headers=['Date','S.No','Customer','Contact','Return Trx','Trx No','Item','Exp Date','Qty','Amount','Reason','Approval','Designation','Limit','Status','Cash/Voucher','Rcv By','CCTV/Time','Trx Time']
    data=[[Paragraph(h,small) for h in headers]]
    for r in rows:
        vals=[r['entry_date'],r['serial_no'],r['customer_name'],r['contact_no'],r['return_trx'],r['trx_no'],r['item_description'],r['item_exp_date'],r['qty'],f"{float(r['total_amount'] or 0):,.2f}",r['reason'],r['approval_name'],r['designation'],('Unlimited' if not r['approval_limit'] else f"{float(r['approval_limit']):,.0f}"),r['system_status'],r['cash_voucher'],r['received_by'],r['cctv_time'],r['trx_time']]
        data.append([Paragraph(str(v or ''),small) for v in vals])
    widths=[45,30,75,55,50,50,105,45,28,55,95,70,62,48,60,55,55,58,48]
    tbl=Table(data,colWidths=widths,repeatRows=1)
    ts=[('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2F6F9F')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),.25,colors.HexColor('#C8D4DD')),('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]
    for i,r in enumerate(rows,1): ts.append(('BACKGROUND',(14,i),(14,i),colors.HexColor('#DDF2E3' if r['system_status']=='Valid Approval' else '#F9DADA')))
    tbl.setStyle(TableStyle(ts)); story.append(tbl)
    doc.build(story); bio.seek(0); return send_file(bio,as_attachment=True,download_name=f'Return_Counter_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf')

def cashier_period_clause(args, alias=""):
    prefix=(alias+".") if alias else ""
    period=(args.get('period') or 'all').strip().lower()
    date=(args.get('date') or '').strip()
    week=(args.get('week') or '').strip()
    month=(args.get('month') or '').strip()
    year=(args.get('year') or '').strip()
    date_from=(args.get('from') or '').strip()
    date_to=(args.get('to') or '').strip()
    clauses=[]; params=[]
    if date_from:
        clauses.append(f"{prefix}closing_date>=?"); params.append(date_from)
    if date_to:
        clauses.append(f"{prefix}closing_date<=?"); params.append(date_to)
    if period=='day' and date:
        clauses.append(f"{prefix}closing_date=?"); params.append(date)
    elif period=='week' and week:
        try:
            y,w=week.split('-W'); d=datetime.fromisocalendar(int(y),int(w),1)
            clauses.append(f"{prefix}closing_date BETWEEN ? AND ?")
            params.extend([d.date().isoformat(),(d+timedelta(days=6)).date().isoformat()])
        except Exception: pass
    elif period=='month' and month:
        clauses.append(f"substr({prefix}closing_date,1,7)=?"); params.append(month)
    elif period=='year' and year:
        clauses.append(f"substr({prefix}closing_date,1,4)=?"); params.append(year)
    return (" AND "+" AND ".join(clauses) if clauses else ""),params,period

@app.get("/api/cashier-closing/dashboard")
@require_permission("cashier_view")
def cashier_closing_dashboard():
    where,params,period=cashier_period_clause(request.args)
    conn=db(); cur=conn.cursor()
    cash_total_expr="(COALESCE(first_total,0)+COALESCE(second_total,0)+COALESCE(third_total,0)+COALESCE(fourth_total,0)+COALESCE(total_closing_cash,0))"
    totals=cur.execute(f"""SELECT
        COALESCE(SUM({cash_total_expr}),0) total_closing_cash,
        COALESCE(SUM(system_total_sale),0) system_sales,
        COALESCE(SUM(collection_difference),0) cash_difference,
        COALESCE(SUM(card_difference),0) card_difference,
        COALESCE(SUM(CASE WHEN collection_difference<0 THEN ABS(collection_difference) ELSE 0 END),0) short_amount,
        COALESCE(SUM(CASE WHEN collection_difference>0 THEN collection_difference ELSE 0 END),0) excess_amount
        FROM cashier_closings WHERE 1=1 {where}""",params).fetchone()
    if period=='week': label="strftime('%Y-W%W',closing_date)"
    elif period=='month': label="substr(closing_date,1,7)"
    elif period=='year': label="substr(closing_date,1,4)"
    else: label="closing_date"
    summary=cur.execute(f"""SELECT {label} period_label,MIN(closing_date) period_start,MAX(closing_date) period_end,
        COALESCE(SUM({cash_total_expr}),0) total_closing_cash,
        COALESCE(SUM(system_total_sale),0) system_sales,COALESCE(SUM(collection_difference),0) cash_difference,
        COALESCE(SUM(ivend_pos),0) ivend_pos,COALESCE(SUM(settlement_bank),0) settlement_bank,
        COALESCE(SUM(card_difference),0) card_difference,
        COALESCE(SUM(CASE WHEN collection_difference<0 THEN ABS(collection_difference) ELSE 0 END),0) short_amount,
        COALESCE(SUM(CASE WHEN collection_difference>0 THEN collection_difference ELSE 0 END),0) excess_amount
        FROM cashier_closings WHERE 1=1 {where} GROUP BY {label} ORDER BY period_start DESC""",params).fetchall()
    conn.close(); return jsonify({"totals":dict(totals),"summary":[dict(r) for r in summary],"period":period})

@app.get("/api/cashier-closing")
@require_permission("cashier_view")
def cashier_closing_list():
    q=request.args.get('q','').strip(); status=request.args.get('status','').strip(); where,params,period=cashier_period_clause(request.args)
    sql="SELECT * FROM cashier_closings WHERE 1=1"+where
    if status:
        sql+=" AND audit_status=?"; params.append(status)
    if q:
        sql+=" AND (employee_id LIKE ? OR employee_name LIKE ? OR audit_status LIKE ? OR card_status LIKE ? OR closing_date LIKE ? OR remarks LIKE ?)"; like=f"%{q}%"; params += [like]*6
    sql+=" ORDER BY closing_date DESC,employee_name,employee_id LIMIT 50000"
    conn=db(); rows=conn.execute(sql,params).fetchall(); conn.close(); return jsonify([dict(r) for r in rows])


@app.get("/api/cashier-closing/note-summary")
@require_permission("cashier_view")
def cashier_closing_note_summary():
    q=request.args.get('q','').strip(); where,params,period=cashier_period_clause(request.args)
    sql="SELECT * FROM cashier_closings WHERE 1=1"+where
    if q:
        sql+=" AND (remarks LIKE ? OR employee_id LIKE ? OR employee_name LIKE ? OR closing_date LIKE ?)"
        like=f"%{q}%"; params += [like]*4
    sql+=" ORDER BY closing_date DESC, employee_name"
    conn=db(); rows=[dict(r) for r in conn.execute(sql,params).fetchall()]; conn.close()
    grouped={}
    for r in rows:
        note=(r.get('remarks') or '').strip() or 'No Note / Remark'
        g=grouped.setdefault(note,{'note':note,'records':0,'cashiers':set(),'short_amount':0.0,'excess_amount':0.0,'first_date':r['closing_date'],'last_date':r['closing_date']})
        g['records']+=1; g['cashiers'].add(str(r.get('employee_id') or ''))
        diff=float(r.get('collection_difference') or 0)
        if diff<0:g['short_amount']+=abs(diff)
        elif diff>0:g['excess_amount']+=diff
        g['first_date']=min(g['first_date'],r['closing_date']);g['last_date']=max(g['last_date'],r['closing_date'])
    summary=[]
    for g in grouped.values():g['cashiers']=len(g['cashiers']);summary.append(g)
    summary.sort(key=lambda x:(-x['records'],x['note'].lower()))
    return jsonify({'summary':summary,'details':rows})

@app.get("/api/cashier-employees")
@require_permission("cashier_view")
def cashier_employees():
    conn=db(); rows=conn.execute("SELECT * FROM cashier_employees ORDER BY employee_name").fetchall(); conn.close(); return jsonify([dict(r) for r in rows])

@app.post("/api/cashier-closing/bulk")
@require_permission("cashier_import")
def cashier_closing_bulk():
    payload=request.get_json(force=True) or {}; entries=payload.get('entries') or []
    if not isinstance(entries,list) or not entries: return jsonify({"error":"At least one closing entry is required"}),400
    if len(entries)>50: return jsonify({"error":"Maximum 50 entries can be saved at one time"}),400
    conn=db(); now=datetime.now().isoformat(timespec='seconds'); inserted=0; updated=0
    cols=['closing_date','employee_id','employee_name','first_5000','first_1000','first_500','first_total','second_5000','second_1000','second_500','second_total','third_5000','third_1000','third_500','third_total','fourth_5000','fourth_1000','fourth_500','fourth_total','close_5000','close_1000','close_500','close_100','close_75','close_50','close_20','close_10','close_5','close_2','close_1','total_closing_cash','system_total_sale','collection_difference','audit_status','remarks','ivend_pos','settlement_bank','card_difference','card_status','card_remarks']
    for e in entries:
        if not str(e.get('employee_id') or '').strip(): continue
        def n(key):
            try: return float(e.get(key) or 0)
            except: return 0.0
        first=n('first_5000')*5000+n('first_1000')*1000+n('first_500')*500
        second=n('second_5000')*5000+n('second_1000')*1000+n('second_500')*500
        third=n('third_5000')*5000+n('third_1000')*1000+n('third_500')*500
        fourth=n('fourth_5000')*5000+n('fourth_1000')*1000+n('fourth_500')*500
        close=n('close_5000')*5000+n('close_1000')*1000+n('close_500')*500+n('close_100')*100+n('close_75')*75+n('close_50')*50+n('close_20')*20+n('close_10')*10+n('close_5')*5+n('close_2')*2+n('close_1')
        total_cash=first+second+third+fourth+close
        diff=total_cash-n('system_total_sale')
        card_diff=n('ivend_pos')-n('settlement_bank')
        e.update({'first_total':first,'second_total':second,'third_total':third,'fourth_total':fourth,'total_closing_cash':close,'collection_difference':diff,'audit_status':'Short' if diff<0 else 'Excess' if diff>0 else 'Matched','card_difference':card_diff,'card_status':'Matched' if abs(card_diff)<0.01 else 'POS Excess' if card_diff>0 else 'POS Short'})
        if not str(e.get('remarks') or '').strip(): e['remarks']=f"Cash shortage Rs. {abs(diff):,.2f}" if diff<0 else f"Cash excess Rs. {diff:,.2f}" if diff>0 else 'Cash matched'
        if not str(e.get('card_remarks') or '').strip(): e['card_remarks']='POS settlement matched' if abs(card_diff)<0.01 else f"POS excess Rs. {card_diff:,.2f}" if card_diff>0 else f"POS short Rs. {abs(card_diff):,.2f}"
        vals=[]
        for c in cols:
            v=e.get(c,'')
            if c not in ('closing_date','employee_id','employee_name','audit_status','remarks','card_status','card_remarks'):
                try:v=float(v or 0)
                except:v=0.0
            vals.append(v)
        existing=conn.execute("SELECT id FROM cashier_closings WHERE closing_date=? AND employee_id=?",(vals[0],str(vals[1]))).fetchone()
        if existing:
            set_sql=','.join(f"{c}=?" for c in cols[2:])
            conn.execute(f"UPDATE cashier_closings SET {set_sql},created_at=? WHERE id=?",tuple(vals[2:])+ (now,existing['id'])); updated+=1
        else:
            conn.execute("INSERT INTO cashier_closings("+','.join(cols)+",created_at) VALUES("+','.join(['?']*(len(cols)+1))+')',tuple(vals)+(now,)); inserted+=1
    conn.commit(); conn.close(); create_automatic_backup(force=True); audit('Bulk Save','Cashier Closing',f'{inserted} inserted, {updated} updated'); return jsonify({'inserted':inserted,'updated':updated})

@app.post("/api/cashier-closing/import")
@require_permission("cashier_import")
def cashier_closing_import():
    file=request.files.get('file')
    if not file:return jsonify({'error':'Excel file required'}),400
    wb=load_workbook(file,read_only=True,data_only=True); conn=db(); now=datetime.now().isoformat(timespec='seconds'); inserted=updated=0
    if 'Employee Database' in wb.sheetnames:
        for row in wb['Employee Database'].iter_rows(min_row=2,values_only=True):
            if row and row[0] not in (None,''):
                conn.execute("INSERT OR REPLACE INTO cashier_employees(employee_id,employee_name) VALUES(?,?)",(str(row[0]).replace('.0',''),str(row[1] or '').strip()))
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit(): continue
        ws=wb[sheet_name]; raw_date=ws.cell(2,1).value
        if not raw_date: continue
        closing_date=raw_date.strftime('%Y-%m-%d') if hasattr(raw_date,'strftime') else str(raw_date)
        for row in ws.iter_rows(min_row=5,max_col=39,values_only=True):
            if row[0] in (None,'','-'): continue
            emp_id=str(row[0]).replace('.0',''); name=str(row[1] or '').strip(); name='Employee Not Found' if name=='#N/A' else name
            numeric=[]
            for v in row[2:32]:
                try:numeric.append(float(v or 0))
                except:numeric.append(0.0)
            diff=float(row[31] or 0); audit_status=str(row[32] or ('Short' if diff<0 else 'Excess' if diff>0 else 'Matched'))
            remarks=str(row[33] or '').strip() or (f"Cash shortage Rs. {abs(diff):,.2f}" if diff<0 else f"Cash excess Rs. {diff:,.2f}" if diff>0 else 'Cash matched')
            ivend=float(row[34] or 0); bank=float(row[35] or 0); card_diff=float(row[36] or (ivend-bank))
            card_status=str(row[37] or ('Matched' if abs(card_diff)<0.01 else 'POS Excess' if card_diff>0 else 'POS Short'))
            card_remarks=str(row[38] or '').strip() or ('POS settlement matched' if abs(card_diff)<0.01 else f"POS excess Rs. {card_diff:,.2f}" if card_diff>0 else f"POS short Rs. {abs(card_diff):,.2f}")
            vals=[closing_date,emp_id,name]+numeric+[audit_status,remarks,ivend,bank,card_diff,card_status,card_remarks]
            existing=conn.execute("SELECT id FROM cashier_closings WHERE closing_date=? AND employee_id=?",(closing_date,emp_id)).fetchone()
            cols=['closing_date','employee_id','employee_name','first_5000','first_1000','first_500','first_total','second_5000','second_1000','second_500','second_total','third_5000','third_1000','third_500','third_total','fourth_5000','fourth_1000','fourth_500','fourth_total','close_5000','close_1000','close_500','close_100','close_75','close_50','close_20','close_10','close_5','close_2','close_1','total_closing_cash','system_total_sale','collection_difference','audit_status','remarks','ivend_pos','settlement_bank','card_difference','card_status','card_remarks']
            if existing:
                conn.execute("UPDATE cashier_closings SET "+','.join(f"{c}=?" for c in cols[2:])+",source_sheet=?,created_at=? WHERE id=?",tuple(vals[2:])+(sheet_name,now,existing['id'])); updated+=1
            else:
                conn.execute("INSERT INTO cashier_closings("+','.join(cols)+",source_sheet,created_at) VALUES("+','.join(['?']*(len(cols)+2))+')',tuple(vals)+(sheet_name,now)); inserted+=1
    conn.commit(); conn.close(); wb.close(); create_automatic_backup(force=True); audit('Import','Cashier Closing',f'{inserted} inserted, {updated} updated'); return jsonify({'inserted':inserted,'updated':updated})

@app.get("/api/export-cashier-closing")
@require_permission("export_data")
def export_cashier_closing():
    where,params,period=cashier_period_clause(request.args)
    q=request.args.get('q','').strip(); status=request.args.get('status','').strip(); note_summary=request.args.get('note_summary','')=='1'
    sql="SELECT * FROM cashier_closings WHERE 1=1"+where
    if status:
        sql+=" AND audit_status=?"; params.append(status)
    if q:
        sql+=" AND (employee_id LIKE ? OR employee_name LIKE ? OR audit_status LIKE ? OR card_status LIKE ? OR closing_date LIKE ?)"
        like=f"%{q}%"; params += [like]*5
    sql+=" ORDER BY closing_date,employee_name,employee_id"
    conn=db(); rows=conn.execute(sql,params).fetchall(); conn.close()
    wb=Workbook(); ws=wb.active; ws.title="Note Wise Report" if note_summary else "Cashier Closing"
    if note_summary:
        grouped={}
        for r in rows:
            note=(r['remarks'] or '').strip() or 'No Note / Remark'
            g=grouped.setdefault(note,{'records':0,'cashiers':set(),'short':0.0,'excess':0.0,'first':r['closing_date'],'last':r['closing_date']})
            g['records']+=1;g['cashiers'].add(str(r['employee_id']))
            diff=float(r['collection_difference'] or 0)
            if diff<0:g['short']+=abs(diff)
            elif diff>0:g['excess']+=diff
            g['first']=min(g['first'],r['closing_date']);g['last']=max(g['last'],r['closing_date'])
        ws.append(['Note / Remark','Records','Cashiers','Shortage Amount','Excess Amount','First Date','Last Date'])
        from openpyxl.styles import Font, PatternFill, Alignment
        for c in ws[1]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='0B3D91');c.alignment=Alignment(horizontal='center')
        for note,g in sorted(grouped.items(),key=lambda x:(-x[1]['records'],x[0].lower())):ws.append([note,g['records'],len(g['cashiers']),g['short'],g['excess'],g['first'],g['last']])
        ws.column_dimensions['A'].width=55
        for col in 'BCDEFG':ws.column_dimensions[col].width=18
        for row in ws.iter_rows(min_row=2):row[3].number_format='#,##0.00';row[4].number_format='#,##0.00'
        ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions

        detail=wb.create_sheet('Collection Denomination Details')
        basic_denom=['5000','1000','500']
        closing_denom=['5000','1000','500','100','75','50','20','10','5','2','1']
        headers=['Note / Remark','Closing Date','Employee ID','Employee Name']
        for label in ['1st','2nd','3rd','4th']:
            headers += [f'{label} {d}' for d in basic_denom] + [f'{label} Total']
        headers += [f'Closing {d}' for d in closing_denom] + ['Closing Total']
        headers += ['Grand Total Cash','System Total Sale','Shortage Amount','Excess Amount','Audit Status']
        detail.append(headers)
        for c in detail[1]:
            c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='0B3D91');c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        for r in rows:
            note=(r['remarks'] or '').strip() or 'No Note / Remark'
            row=[note,r['closing_date'],r['employee_id'],r['employee_name']]
            for prefix in ['first','second','third','fourth']:
                row += [float(r[f'{prefix}_{d}'] or 0) for d in basic_denom]
                row += [float(r[f'{prefix}_total'] or 0)]
            row += [float(r[f'close_{d}'] or 0) for d in closing_denom]
            row += [float(r['total_closing_cash'] or 0)]
            grand=sum(float(r[f'{p}_total'] or 0) for p in ['first','second','third','fourth'])+float(r['total_closing_cash'] or 0)
            diff=float(r['collection_difference'] or 0)
            row += [grand,float(r['system_total_sale'] or 0),abs(diff) if diff<0 else 0,diff if diff>0 else 0,r['audit_status']]
            detail.append(row)
        detail.freeze_panes='E2';detail.auto_filter.ref=detail.dimensions
        detail.column_dimensions['A'].width=45;detail.column_dimensions['B'].width=14;detail.column_dimensions['C'].width=13;detail.column_dimensions['D'].width=25
        for col in range(5,detail.max_column+1):detail.column_dimensions[detail.cell(1,col).column_letter].width=12
        for row in detail.iter_rows(min_row=2):
            for c in row[4:-1]:c.number_format='#,##0.00'
        bio=io.BytesIO();wb.save(bio);bio.seek(0)
        return send_file(bio,as_attachment=True,download_name='Cashier_Note_Wise_Denomination_Report.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    headers=["Closing Date","Employee ID","Employee Name","1st Collection","2nd Collection","3rd Collection","4th Collection","Closing Cash","Total Cash Collected","System Total Sale","Shortage Amount","Excess Amount","Audit Status","Note / Remarks","iVend POS","Settlement Bank","Card Difference","Card Status","Card Remarks"]
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in ws[1]:
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="0B3D91"); c.alignment=Alignment(horizontal="center")
    cash_total=lambda r: sum(float(r[k] or 0) for k in ('first_total','second_total','third_total','fourth_total','total_closing_cash'))
    for r in rows:
        diff=float(r['collection_difference'] or 0); short_amt=abs(diff) if diff<0 else 0; excess_amt=diff if diff>0 else 0
        ws.append([r['closing_date'],r['employee_id'],r['employee_name'],r['first_total'],r['second_total'],r['third_total'],r['fourth_total'],r['total_closing_cash'],cash_total(r),r['system_total_sale'],short_amt,excess_amt,r['audit_status'],r['remarks'],r['ivend_pos'],r['settlement_bank'],r['card_difference'],r['card_status'],r['card_remarks']])
    widths={'A':14,'B':13,'C':25,'D':16,'E':16,'F':16,'G':16,'H':16,'I':19,'J':19,'K':17,'L':17,'M':14,'N':32,'O':17,'P':20,'Q':17,'R':15,'S':30}
    for col,width in widths.items(): ws.column_dimensions[col].width=width
    for row in ws.iter_rows(min_row=2):
        for c in list(row[3:12])+list(row[14:17]): c.number_format='#,##0.00;[Red]-#,##0.00'
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name='Cashier_Closing_Filtered.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get("/api/pdf-cashier-closing")
@require_permission("export_data")
def pdf_cashier_closing():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm
    where,params,period=cashier_period_clause(request.args)
    q=request.args.get('q','').strip(); status=request.args.get('status','').strip()
    sql="SELECT * FROM cashier_closings WHERE 1=1"+where
    if status:
        sql+=" AND audit_status=?"; params.append(status)
    if q:
        sql+=" AND (employee_id LIKE ? OR employee_name LIKE ? OR audit_status LIKE ? OR card_status LIKE ? OR closing_date LIKE ?)"
        like=f"%{q}%"; params += [like]*5
    sql+=" ORDER BY closing_date,employee_name,employee_id"
    conn=db(); rows=conn.execute(sql,params).fetchall(); conn.close()
    cash_total=lambda r: sum(float(r[k] or 0) for k in ('first_total','second_total','third_total','fourth_total','total_closing_cash'))
    total_cash=sum(cash_total(r) for r in rows); total_sales=sum(float(r['system_total_sale'] or 0) for r in rows)
    cash_diff=sum(float(r['collection_difference'] or 0) for r in rows)
    short_amt=sum(abs(float(r['collection_difference'] or 0)) for r in rows if float(r['collection_difference'] or 0)<0)
    excess_amt=sum(float(r['collection_difference'] or 0) for r in rows if float(r['collection_difference'] or 0)>0)
    bio=io.BytesIO(); doc=SimpleDocTemplate(bio,pagesize=landscape(A3),rightMargin=8*mm,leftMargin=8*mm,topMargin=8*mm,bottomMargin=8*mm)
    styles=getSampleStyleSheet(); story=[Paragraph('Cashier Closing Report',styles['Title']),Spacer(1,4*mm)]
    summary=[["Total Closing Cash",f"Rs. {total_cash:,.2f}","System Sales",f"Rs. {total_sales:,.2f}"],["Cash Difference",f"Rs. {cash_diff:,.2f}","Short Amount",f"Rs. {short_amt:,.2f}"],["Excess Amount",f"Rs. {excess_amt:,.2f}","Filtered Rows",str(len(rows))]]
    st=Table(summary,colWidths=[38*mm,42*mm,38*mm,42*mm]); st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#E8F1FB')),('GRID',(0,0),(-1,-1),0.4,colors.grey),('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),('ALIGN',(1,0),(1,-1),'RIGHT'),('ALIGN',(3,0),(3,-1),'RIGHT'),('BOTTOMPADDING',(0,0),(-1,-1),6),('TOPPADDING',(0,0),(-1,-1),6)])); story += [st,Spacer(1,5*mm)]
    data=[["Date","Emp ID","Employee Name","1st","2nd","3rd","4th","Closing","Total Cash","System Sale","Shortage","Excess","Status","Remarks","Card Diff"]]
    for r in rows:
        diff=float(r['collection_difference'] or 0); short_amt=abs(diff) if diff<0 else 0; excess_amt=diff if diff>0 else 0
        data.append([r['closing_date'],r['employee_id'],Paragraph(str(r['employee_name'] or ''),styles['BodyText']),f"{float(r['first_total'] or 0):,.2f}",f"{float(r['second_total'] or 0):,.2f}",f"{float(r['third_total'] or 0):,.2f}",f"{float(r['fourth_total'] or 0):,.2f}",f"{float(r['total_closing_cash'] or 0):,.2f}",f"{cash_total(r):,.2f}",f"{float(r['system_total_sale'] or 0):,.2f}",f"{short_amt:,.2f}",f"{excess_amt:,.2f}",r['audit_status'],Paragraph(str(r['remarks'] or ''),styles['BodyText']),f"{float(r['card_difference'] or 0):,.2f}"])
    table=Table(data,repeatRows=1,colWidths=[20*mm,15*mm,34*mm,20*mm,20*mm,20*mm,20*mm,22*mm,24*mm,24*mm,21*mm,21*mm,17*mm,34*mm,20*mm])
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0B3D91')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#B8C7D9')),('ALIGN',(3,1),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F4F7FA')])]))
    story.append(table); doc.build(story); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name='Cashier_Closing_Filtered.pdf',mimetype='application/pdf')

@app.get("/api/vendors")
@require_permission("vendor_view")
def vendors():
    conn=db()
    rows=conn.execute("SELECT * FROM vendors ORDER BY vendor_code").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.post("/api/vendors")
@require_permission("vendor_add")
def add_vendor():
    data=request.get_json(force=True)
    conn=db()
    conn.execute("""INSERT INTO vendors(vendor_code,vendor_name,security_deposit,advance,phone,description,status)
                    VALUES(?,?,?,?,?,?,?)""",(
        data.get("vendor_code"),data.get("vendor_name"),float(data.get("security_deposit") or 0),
        float(data.get("advance") or 0),data.get("phone",""),data.get("description",""),
        data.get("status","Active")
    ))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.put("/api/vendors/<int:vendor_id>")
@require_permission("vendor_edit")
def update_vendor(vendor_id):
    data=request.get_json(force=True); conn=db()
    row=conn.execute("SELECT * FROM vendors WHERE id=?",(vendor_id,)).fetchone()
    if not row: conn.close(); return jsonify({"error":"Vendor not found"}),404
    conn.execute("""UPDATE vendors SET vendor_code=?,vendor_name=?,security_deposit=?,advance=?,phone=?,description=?,status=? WHERE id=?""",(
        data.get("vendor_code"),data.get("vendor_name"),float(data.get("security_deposit") or 0),float(data.get("advance") or 0),
        data.get("phone",""),data.get("description",""),data.get("status","Active"),vendor_id))
    conn.commit(); conn.close(); audit("Update","Vendor",f"Vendor ID {vendor_id}")
    return jsonify({"ok":True})

@app.delete("/api/vendors/<int:vendor_id>")
@require_permission("vendor_delete")
def delete_vendor(vendor_id):
    conn=db(); row=conn.execute("SELECT vendor_code,vendor_name FROM vendors WHERE id=?",(vendor_id,)).fetchone()
    if not row: conn.close(); return jsonify({"error":"Vendor not found"}),404
    tx_count=conn.execute("SELECT COUNT(*) FROM vendor_ledger WHERE vendor_code=?",(row["vendor_code"],)).fetchone()[0]
    if tx_count:
        conn.execute("UPDATE vendors SET status='Inactive' WHERE id=?",(vendor_id,)); mode="inactivated"
    else:
        conn.execute("DELETE FROM vendors WHERE id=?",(vendor_id,)); mode="deleted"
    conn.commit(); conn.close(); audit(mode.title(),"Vendor",f"{row['vendor_code']} - {row['vendor_name']}")
    return jsonify({"ok":True,"mode":mode,"transactions":tx_count})

@app.get("/api/vendor-ledger")
@require_permission("ledger_view")
def vendor_ledger():
    code=request.args.get("code","")
    q=request.args.get("q","")
    sql="SELECT * FROM vendor_ledger WHERE 1=1"; params=[]
    if code: sql+=" AND vendor_code=?"; params.append(code)
    if q:
        sql+=" AND (vendor_code LIKE ? OR vendor_name LIKE ? OR document_number LIKE ? OR description LIKE ? OR gate_pass LIKE ? OR bilty_no LIKE ?)"
        like=f"%{q}%"; params += [like]*6
    sql+=" ORDER BY tx_date DESC,id DESC LIMIT 5000"
    conn=db(); rows=conn.execute(sql,params).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.post("/api/vendor-ledger")
@require_permission("ledger_add")
def add_vendor_tx():
    data=request.get_json(force=True)
    conn=db()
    conn.execute("""INSERT INTO vendor_ledger(
        vendor_code,vendor_name,tx_date,tx_type,description,document_number,debit,credit,payment_status,
        gate_pass,bilty_no,created_at
    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",(
        data.get("vendor_code"),data.get("vendor_name"),data.get("tx_date"),
        data.get("tx_type"),data.get("description",""),str(data.get("document_number") or ""),float(data.get("debit") or 0),
        float(data.get("credit") or 0),data.get("payment_status","Pending"),
        data.get("gate_pass",""),data.get("bilty_no",""),datetime.now().isoformat(timespec="seconds")
    ))
    conn.commit(); conn.close()
    return jsonify({"ok":True})


@app.post("/api/vendor-ledger/bulk")
@require_permission("ledger_add")
def bulk_vendor_ledger():
    data = request.get_json(force=True) or {}
    entries = data.get("entries", [])
    if not isinstance(entries, list) or not entries:
        return jsonify({"error":"No entries supplied"}), 400
    conn = db(); cur = conn.cursor()
    vendors = {str(r["vendor_code"]).strip(): r["vendor_name"] for r in cur.execute("SELECT vendor_code,vendor_name FROM vendors")}
    inserted = 0; skipped = []
    today = datetime.now().strftime("%Y-%m-%d")
    for i, row in enumerate(entries, start=1):
        code = str(row.get("code") or "").strip()
        typed_name = str(row.get("name") or "").strip()
        if not code:
            skipped.append({"row":i,"code":"","reason":"Vendor code is required"}); continue
        name = vendors.get(code)
        if not name:
            if not typed_name:
                skipped.append({"row":i,"code":code,"reason":"Vendor not found and name is blank"}); continue
            cur.execute("""INSERT OR IGNORE INTO vendors(vendor_code,vendor_name,status) VALUES(?,?, 'Active')""",(code,typed_name))
            name = typed_name; vendors[code] = name
        debit = n(row.get("debit")); credit = n(row.get("credit"))
        if not row.get("description") and debit == 0 and credit == 0:
            continue
        cur.execute("""INSERT INTO vendor_ledger(
            vendor_code,vendor_name,tx_date,tx_type,description,document_number,debit,credit,payment_status,
            gate_pass,bilty_no,created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",(
            code,name,str(row.get("date") or today),"Manual Bulk",str(row.get("description") or ""),
            str(row.get("document_number") or ""),debit,credit,str(row.get("status") or "Pending"),"","",datetime.now().isoformat(timespec="seconds")
        ))
        inserted += 1
    conn.commit(); conn.close()
    audit("Bulk Add","Vendor Ledger",f"Inserted {inserted}; skipped {len(skipped)}")
    return jsonify({"ok":True,"inserted":inserted,"skipped":skipped})


@app.post("/api/vendor-ledger/import-large")
@require_permission("ledger_add")
def import_vendor_ledger_large():
    """Import up to 50,000 Excel/CSV rows using columns A:H."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error":"Please select an Excel or CSV file"}), 400
    suffix = Path(f.filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".csv"}:
        return jsonify({"error":"Only .xlsx, .xlsm or .csv files are supported"}), 400

    max_rows = 50000
    parsed = []
    try:
        if suffix == ".csv":
            import csv
            stream = io.StringIO(f.stream.read().decode("utf-8-sig", errors="replace"))
            reader = csv.reader(stream)
            for idx, row in enumerate(reader):
                if idx == 0 and row and "vendor" in str(row[0]).lower():
                    continue
                if idx >= max_rows + 1:
                    break
                parsed.append(row[:8])
        else:
            wb = load_workbook(f.stream, read_only=True, data_only=True)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=8, values_only=True)):
                vals = list(row)
                if idx == 0 and vals and "vendor" in str(vals[0]).lower():
                    continue
                if len(parsed) >= max_rows:
                    break
                parsed.append(vals)
            wb.close()
    except Exception as exc:
        return jsonify({"error":f"Could not read file: {exc}"}), 400

    conn = db(); cur = conn.cursor()
    vendors = {str(r["vendor_code"]).strip(): r["vendor_name"] for r in cur.execute("SELECT vendor_code,vendor_name FROM vendors")}
    inserted = 0; skipped = 0
    created_at = datetime.now().isoformat(timespec="seconds")

    def clean(v):
        if v is None: return ""
        if isinstance(v, float) and v.is_integer(): return str(int(v))
        return str(v).strip()
    def date_text(v):
        if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
        if isinstance(v, (int,float)):
            try: return (datetime(1899,12,30)+timedelta(days=float(v))).strftime("%Y-%m-%d")
            except Exception: return clean(v)
        txt=clean(v)
        for fmt in ("%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%m/%d/%Y"):
            try: return datetime.strptime(txt,fmt).strftime("%Y-%m-%d")
            except Exception: pass
        return txt

    batch=[]
    for row in parsed[:max_rows]:
        row = list(row) + [None] * (8-len(row))
        code,name_typed,doc_no,post_date,desc,debit_v,credit_v,_balance = row[:8]
        code=clean(code); name_typed=clean(name_typed)
        if not code:
            skipped += 1; continue
        name=vendors.get(code)
        if not name:
            if not name_typed:
                skipped += 1; continue
            cur.execute("INSERT OR IGNORE INTO vendors(vendor_code,vendor_name,status) VALUES(?,?, 'Active')",(code,name_typed))
            name=name_typed; vendors[code]=name
        debit=n(debit_v); credit=n(credit_v)
        if not clean(doc_no) and not clean(desc) and debit==0 and credit==0:
            skipped += 1; continue
        batch.append((code,name,date_text(post_date),"Excel Import",clean(desc),clean(doc_no),debit,credit,"Pending","","",created_at))
        if len(batch)>=1000:
            cur.executemany("""INSERT INTO vendor_ledger(vendor_code,vendor_name,tx_date,tx_type,description,document_number,debit,credit,payment_status,gate_pass,bilty_no,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",batch)
            inserted += len(batch); batch=[]
    if batch:
        cur.executemany("""INSERT INTO vendor_ledger(vendor_code,vendor_name,tx_date,tx_type,description,document_number,debit,credit,payment_status,gate_pass,bilty_no,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",batch)
        inserted += len(batch)
    conn.commit(); conn.close()
    audit("Large Import","Vendor Ledger",f"Inserted {inserted}; skipped {skipped}; file {f.filename}")
    return jsonify({"ok":True,"inserted":inserted,"skipped":skipped,"limit":max_rows})

@app.delete("/api/vendor-ledger/<int:tx_id>")
def delete_vendor_tx(tx_id):
    conn=db(); conn.execute("DELETE FROM vendor_ledger WHERE id=?",(tx_id,)); conn.commit(); conn.close()
    return jsonify({"ok":True})


@app.post("/api/customer-ledger/delete-selected")
def delete_customer_selected():
    ids = request.get_json(force=True).get("ids", [])
    if not ids:
        return jsonify({"ok": True, "deleted": 0})
    conn = db()
    q = ",".join("?" for _ in ids)
    cur = conn.execute(f"DELETE FROM customer_ledger WHERE id IN ({q})", ids)
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})

@app.delete("/api/customer-ledger/delete-all")
def delete_customer_all():
    conn = db()
    cur = conn.execute("DELETE FROM customer_ledger")
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})

@app.post("/api/duplicates/delete-selected")
def delete_duplicate_selected():
    ids = request.get_json(force=True).get("ids", [])
    if not ids:
        return jsonify({"ok": True, "deleted": 0})
    conn = db()
    q = ",".join("?" for _ in ids)
    cur = conn.execute(f"DELETE FROM duplicate_documents WHERE id IN ({q})", ids)
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})

@app.delete("/api/duplicates/delete-all")
def delete_duplicate_all():
    conn = db()
    cur = conn.execute("DELETE FROM duplicate_documents")
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})

@app.post("/api/vendor-ledger/delete-selected")
def delete_vendor_ledger_selected():
    ids = request.get_json(force=True).get("ids", [])
    if not ids:
        return jsonify({"ok": True, "deleted": 0})
    conn = db()
    q = ",".join("?" for _ in ids)
    cur = conn.execute(f"DELETE FROM vendor_ledger WHERE id IN ({q})", ids)
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})

@app.delete("/api/vendor-ledger/delete-all")
def delete_vendor_ledger_all():
    conn = db()
    cur = conn.execute("DELETE FROM vendor_ledger")
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})


@app.get("/api/export/<module>")
def export(module):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    conn = db()

    if module == "customers":
        ws.title = "Corporate Ledger"
        headers = ["Document No","Posting Date","Description","Debit Amount","Credit Amount","Balance","Customer Code","Customer Name"]
        ws.append(["RAHAT CORPORATE MANAGEMENT DASHBOARD"])
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
        ws.append(headers)
        balance_by_customer = {}
        rows = conn.execute("""SELECT * FROM customer_ledger
                             ORDER BY customer_code,posting_date,document_number""").fetchall()
        for r in rows:
            code = r["customer_code"]
            balance_by_customer[code] = balance_by_customer.get(code,0) + float(r["debit"] or 0) - float(r["credit"] or 0)
            ws.append([r["document_number"],r["posting_date"],r["text"],r["debit"],r["credit"],
                       balance_by_customer[code],code,r["customer_name"]])
    elif module == "duplicates":
        ws.title = "Duplicate Documents"
        headers = ["Document No","Posting Date","Description","Debit Amount","Credit Amount","Balance","Customer Code","Customer Name","Occurrences"]
        ws.append(["DUPLICATE DOCUMENT AUDIT REPORT"])
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
        ws.append(headers)
        rows = conn.execute("SELECT * FROM duplicate_documents ORDER BY document_number,posting_date").fetchall()
        balance = 0
        for r in rows:
            balance += float(r["debit"] or 0) - float(r["credit"] or 0)
            ws.append([r["document_number"],r["posting_date"],r["text"],r["debit"],r["credit"],
                       balance,r["customer_code"],r["customer_name"],r["duplicate_occurrences"]])
    else:
        ws.title = "Scrap Vendor Ledger"
        headers = ["Vendor Code","Vendor Name","Document No.","Posting Date","Description","Debit Amount","Credit Amount","Balance","Payment Status"]
        ws.append(["SCRAP VENDOR LEDGER REPORT"])
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
        ws.append(headers)
        rows = conn.execute("SELECT * FROM vendor_ledger ORDER BY vendor_code,tx_date,id").fetchall()
        balances = {}
        for r in rows:
            code = r["vendor_code"]
            balances[code] = balances.get(code,0) + float(r["debit"] or 0) - float(r["credit"] or 0)
            ws.append([code,r["vendor_name"],r["document_number"],r["tx_date"],r["description"],
                       r["debit"],r["credit"],balances[code],r["payment_status"]])
    conn.close()

    dark = "17365D"
    blue = "0A6ED1"
    gold = "F0AB00"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E1E8")

    ws["A1"].font = Font(bold=True,size=16,color=white)
    ws["A1"].fill = PatternFill("solid",fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    for cell in ws[2]:
        cell.font = Font(bold=True,color=white)
        cell.fill = PatternFill("solid",fgColor=blue)
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border = Border(top=thin,bottom=thin,left=thin,right=thin)
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = Border(bottom=thin)
            if cell.column in (4,5,6,8):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    widths = [18,14,38,16,16,16,16,38,16]
    for i,w in enumerate(widths[:ws.max_column],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio,as_attachment=True,download_name=f"{module}_professional_report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/pdf/<module>")
def export_pdf(module):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm

    conn = db()
    if module == "customers":
        title = "Rahat Corporate Customer Ledger"
        headers = ["Document No","Posting Date","Description","Debit","Credit","Balance"]
        rows = conn.execute("""SELECT document_number,posting_date,text,debit,credit,customer_code
                              FROM customer_ledger ORDER BY customer_code,posting_date,document_number""").fetchall()
        balances = {}
        data = []
        for r in rows:
            code = r["customer_code"]
            balances[code] = balances.get(code,0)+float(r["debit"] or 0)-float(r["credit"] or 0)
            data.append([r["document_number"],r["posting_date"],str(r["text"])[:42],
                         f'{float(r["debit"] or 0):,.2f}',f'{float(r["credit"] or 0):,.2f}',
                         f'{balances[code]:,.2f}'])
    elif module == "duplicates":
        title = "Duplicate Document Audit Report"
        headers = ["Document No","Posting Date","Description","Debit","Credit","Occurrences"]
        rows = conn.execute("""SELECT document_number,posting_date,text,debit,credit,duplicate_occurrences
                              FROM duplicate_documents ORDER BY document_number,posting_date""").fetchall()
        data = [[r["document_number"],r["posting_date"],str(r["text"])[:42],
                 f'{float(r["debit"] or 0):,.2f}',f'{float(r["credit"] or 0):,.2f}',
                 r["duplicate_occurrences"]] for r in rows]
    else:
        title = "Scrap Vendor Ledger Report"
        headers = ["Date","Vendor","Description","Debit","Credit","Balance"]
        rows = conn.execute("""SELECT tx_date,vendor_name,description,debit,credit,vendor_code
                              FROM vendor_ledger ORDER BY vendor_code,tx_date,id""").fetchall()
        balances={}
        data=[]
        for r in rows:
            code=r["vendor_code"]
            balances[code]=balances.get(code,0)+float(r["debit"] or 0)-float(r["credit"] or 0)
            data.append([r["tx_date"],r["vendor_name"],str(r["description"])[:42],
                         f'{float(r["debit"] or 0):,.2f}',f'{float(r["credit"] or 0):,.2f}',
                         f'{balances[code]:,.2f}'])
    conn.close()

    bio=io.BytesIO()
    doc=SimpleDocTemplate(bio,pagesize=landscape(A3),rightMargin=10*mm,leftMargin=10*mm,topMargin=10*mm,bottomMargin=10*mm)
    styles=getSampleStyleSheet()
    story=[Paragraph(title,styles["Title"]),Spacer(1,6)]
    table=Table([headers]+data,repeatRows=1,colWidths=[28*mm,26*mm,82*mm,27*mm,27*mm,28*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#17365D")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#D9E1E8")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F7FA")]),
        ("ALIGN",(3,1),(-1,-1),"RIGHT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)
    bio.seek(0)
    return send_file(bio,as_attachment=True,download_name=f"{module}_professional_report.pdf",
                     mimetype="application/pdf")

@app.get("/api/users")
@require_permission("user_manage")
def list_users():
    conn=auth_db(); rows=conn.execute("SELECT id,full_name,username,role_name,user_type,store_access,permissions,status,created_at,last_login FROM users ORDER BY id").fetchall(); conn.close()
    out=[]
    for r in rows:
        d=dict(r)
        try: d["permissions"]=json.loads(d["permissions"] or "[]")
        except Exception: d["permissions"]=[]
        out.append(d)
    return jsonify(out)

@app.post("/api/users")
@require_permission("user_manage")
def create_user():
    data=request.get_json(force=True)
    if not data.get("username") or not data.get("password") or not data.get("full_name"):
        return jsonify({"error":"Full name, username and password are required"}),400
    conn=auth_db()
    try:
        conn.execute("""INSERT INTO users(full_name,username,password_hash,role_name,user_type,store_access,permissions,status,created_at)
                      VALUES(?,?,?,?,?,?,?,?,?)""",(
            data["full_name"].strip(),data["username"].strip(),generate_password_hash(data["password"]),
            data.get("role_name","Local User"),data.get("user_type","Local"),data.get("store_access","ALL"),
            json.dumps(data.get("permissions",[])),data.get("status","Active"),datetime.now().isoformat(timespec="seconds")))
        conn.commit()
        backup_users(conn)
    except sqlite3.IntegrityError:
        conn.close(); return jsonify({"error":"Username already exists"}),409
    conn.close(); audit("Create","User",data["username"]); return jsonify({"ok":True})

@app.put("/api/users/<int:user_id>")
@require_permission("user_manage")
def update_user(user_id):
    data=request.get_json(force=True); conn=auth_db(); row=conn.execute("SELECT * FROM users WHERE id=?",(user_id,)).fetchone()
    if not row: conn.close(); return jsonify({"error":"User not found"}),404
    if str(row["username"] or "").strip().lower()=="rahat":
        data["role_name"]="Super Admin"; data["status"]="Active"; data["permissions"]=PERMISSIONS
    if row["role_name"]=="Super Admin" and user_id!=g.user["id"] and data.get("status")=="Blocked":
        conn.close(); return jsonify({"error":"Super Admin cannot be blocked"}),400
    fields=[data.get("full_name",row["full_name"]),data.get("username",row["username"]),data.get("role_name",row["role_name"]),
            data.get("user_type",row["user_type"]),data.get("store_access",row["store_access"]),json.dumps(data.get("permissions",json.loads(row["permissions"] or "[]"))),
            data.get("status",row["status"]),user_id]
    try:
        conn.execute("UPDATE users SET full_name=?,username=?,role_name=?,user_type=?,store_access=?,permissions=?,status=? WHERE id=?",fields)
        if data.get("password"): conn.execute("UPDATE users SET password_hash=? WHERE id=?",(generate_password_hash(data["password"]),user_id))
        conn.commit()
        backup_users(conn)
    except sqlite3.IntegrityError:
        conn.close(); return jsonify({"error":"Username already exists"}),409
    conn.close(); audit("Update","User",str(user_id)); return jsonify({"ok":True})

@app.delete("/api/users/<int:user_id>")
@require_permission("user_manage")
def delete_user(user_id):
    if user_id==g.user["id"]: return jsonify({"error":"You cannot delete your own account"}),400
    conn=auth_db(); row=conn.execute("SELECT username,role_name FROM users WHERE id=?",(user_id,)).fetchone()
    if not row: conn.close(); return jsonify({"error":"User not found"}),404
    if row["role_name"]=="Super Admin" or str(row["username"] or "").strip().lower()=="rahat": conn.close(); return jsonify({"error":"Rahat/Admin cannot be deleted"}),400
    conn.close()
    return jsonify({"error":"User deletion is permanently disabled. Block the user instead so the account and rights remain safe."}),400

@app.get("/api/audit-log")
@require_permission("audit_view")
def get_audit_log():
    conn=db(); rows=conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 1000").fetchall(); conn.close(); return jsonify([dict(r) for r in rows])

@app.get("/api/network")
def network():
    try:
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); s.connect(("8.8.8.8",80))
        ip=s.getsockname()[0]; s.close()
    except Exception:
        ip="YOUR-PC-IP"
    return jsonify({"url":f"http://{ip}:5055"})

@app.get("/api/backup")
@require_permission("backup_restore")
def backup():
    create_automatic_backup(force=True)
    name = f"Rahat_Corporate_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    audit("Download", "Backup", name)
    return send_file(DB, as_attachment=True, download_name=name)

@app.get("/api/backup-status")
@require_permission("backup_restore")
def backup_status():
    backups = sorted(BACKUP_DIR.glob("*.db"), key=lambda x: x.stat().st_mtime, reverse=True)
    latest = backups[0] if backups else None
    return jsonify({
        "database_location": str(DB),
        "latest_backup": latest.name if latest else "No backup yet",
        "latest_backup_time": datetime.fromtimestamp(latest.stat().st_mtime).isoformat(timespec="seconds") if latest else "",
        "backup_count": len(backups)
    })

@app.post("/api/restore")
@require_permission("backup_restore")
def restore():
    file = request.files.get("file")
    if not file:
        return jsonify({"error":"No backup selected"}),400
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".db", ".sqlite", ".sqlite3"):
        return jsonify({"error":"Please select a valid .db backup file"}),400
    with tempfile.NamedTemporaryFile(delete=False, suffix=".db", dir=APP_DATA) as tmp:
        temp_path = Path(tmp.name)
        file.save(tmp)
    try:
        test = sqlite3.connect(temp_path)
        result = test.execute("PRAGMA integrity_check").fetchone()[0]
        tables = {r[0] for r in test.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        test.close()
        required = {"customers", "customer_ledger", "vendors", "vendor_ledger", "users"}
        if result != "ok" or not required.issubset(tables):
            return jsonify({"error":"Backup file is invalid or belongs to another software"}),400
        create_automatic_backup(force=True)
        restore_point = BACKUP_DIR / f"before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        if DB.exists(): shutil.copy2(DB, restore_point)
        os.replace(temp_path, DB)
        init_db()
        audit("Restore", "Backup", file.filename or "backup.db")
        return jsonify({"ok":True,"message":"Backup restored successfully. Please login again."})
    finally:
        if temp_path.exists():
            try: temp_path.unlink()
            except OSError: pass



def _excel_rows(file_storage):
    """Read the first worksheet and return normalized dictionaries."""
    data=file_storage.read()
    wb=load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True)
    headers=[]
    for row in rows:
        vals=[str(v).strip() if v is not None else '' for v in row]
        if any(vals):
            headers=[v.lower().replace('\n',' ').strip() for v in vals]
            break
    out=[]
    for row in rows:
        vals=list(row)
        if not any(v not in (None,'') for v in vals): continue
        out.append({headers[i] if i < len(headers) and headers[i] else f'col_{i+1}': vals[i] for i in range(len(vals))})
    wb.close()
    return out

def _pick(row, names):
    for key,val in row.items():
        k=' '.join(str(key).lower().replace('_',' ').replace('.',' ').split())
        if any(n in k for n in names) and val not in (None,''):
            return val
    return ''

def _num(value):
    try:
        if value in (None,''): return 0.0
        return float(str(value).replace(',','').replace('(','-').replace(')','').strip())
    except Exception:
        return 0.0

def _extract_doc_amount(rows):
    result={}
    for r in rows:
        doc=str(_pick(r,['document no','document number','doc no','reference','transaction no','trx no'])).strip()
        if not doc: continue
        amount=_num(_pick(r,['net amount','amount','debit amount','debit']))
        if not amount:
            amount=_num(_pick(r,['credit amount','credit']))
        result.setdefault(doc,0.0)
        result[doc]+=amount
    return result

@app.post('/api/accounts/reconcile')
@require_permission('ledger_view')
def accounts_reconcile():
    source=request.files.get('source_file'); target=request.files.get('target_file')
    if not source or not target: return jsonify({'error':'Select both SAP/Excel and software files'}),400
    a=_extract_doc_amount(_excel_rows(source)); b=_extract_doc_amount(_excel_rows(target))
    conn=db(); now=datetime.now().isoformat(timespec='seconds')
    cur=conn.execute("INSERT INTO reconciliation_runs(run_type,title,created_by,created_at) VALUES(?,?,?,?)",('Ledger',request.form.get('title') or 'SAP vs Software',g.user.get('username'),now))
    run_id=cur.lastrowid; counts={'Matched':0,'Amount Difference':0,'Unmatched':0}
    for doc in sorted(set(a)|set(b)):
        av=float(a.get(doc,0)); bv=float(b.get(doc,0)); diff=round(av-bv,2)
        if doc in a and doc in b and abs(diff)<0.01: status='Matched'
        elif doc in a and doc in b: status='Amount Difference'
        else: status='Unmatched'
        counts[status]+=1
        conn.execute("INSERT INTO reconciliation_results(run_id,source_name,document_number,source_amount,target_amount,difference,status,details) VALUES(?,?,?,?,?,?,?,?)",(run_id,'SAP vs Software',doc,av,bv,diff,status,''))
    conn.commit(); conn.close(); audit('Create','Reconciliation',f'Run {run_id}: {counts}')
    return jsonify({'ok':True,'run_id':run_id,'counts':counts})

@app.get('/api/accounts/reconciliation-results')
@require_permission('ledger_view')
def reconciliation_results():
    conn=db(); run_id=request.args.get('run_id')
    if not run_id:
        row=conn.execute("SELECT id FROM reconciliation_runs ORDER BY id DESC LIMIT 1").fetchone(); run_id=row['id'] if row else 0
    rows=conn.execute("SELECT * FROM reconciliation_results WHERE run_id=? ORDER BY CASE status WHEN 'Amount Difference' THEN 1 WHEN 'Unmatched' THEN 2 ELSE 3 END, document_number",(run_id,)).fetchall() if run_id else []
    conn.close(); return jsonify([dict(r) for r in rows])

@app.get('/api/accounts/aging')
@require_permission('ledger_view')
def accounts_aging():
    kind=(request.args.get('kind') or 'customer').lower(); today=datetime.now().date(); conn=db()
    if kind=='vendor':
        rows=conn.execute("SELECT vendor_code code,vendor_name name,tx_date dt,SUM(debit-credit) amount FROM vendor_ledger GROUP BY vendor_code,vendor_name,tx_date").fetchall()
    else:
        rows=conn.execute("SELECT customer_code code,customer_name name,posting_date dt,SUM(debit-credit) amount FROM customer_ledger GROUP BY customer_code,customer_name,posting_date").fetchall()
    grouped={}
    for r in rows:
        key=(r['code'] or '',r['name'] or '')
        g1=grouped.setdefault(key,{'code':key[0],'name':key[1],'0_30':0,'31_60':0,'61_90':0,'90_plus':0,'total':0,'last_date':''})
        amt=float(r['amount'] or 0); g1['total']+=amt
        try: d=datetime.fromisoformat(str(r['dt'])[:10]).date(); age=max(0,(today-d).days)
        except Exception: age=9999; d=None
        bucket='0_30' if age<=30 else '31_60' if age<=60 else '61_90' if age<=90 else '90_plus'; g1[bucket]+=amt
        if d and str(d)>g1['last_date']: g1['last_date']=str(d)
    conn.close(); return jsonify(sorted(grouped.values(),key=lambda x:abs(x['total']),reverse=True))

@app.get('/api/accounts/exceptions')
@require_permission('ledger_view')
def accounts_exceptions():
    conn=db(); items=[]
    dup=conn.execute("SELECT COUNT(*) c FROM duplicate_documents").fetchone()['c']
    if dup: items.append({'severity':'Critical','type':'Duplicate Documents','count':dup,'details':'Duplicate document records require review.'})
    dif=conn.execute("SELECT COUNT(*) c,COALESCE(SUM(ABS(difference)),0) amount FROM reconciliation_results WHERE status='Amount Difference'").fetchone()
    if dif['c']: items.append({'severity':'Critical','type':'Amount Differences','count':dif['c'],'amount':dif['amount'],'details':'Same document number but different amount.'})
    unm=conn.execute("SELECT COUNT(*) c FROM reconciliation_results WHERE status='Unmatched'").fetchone()['c']
    if unm: items.append({'severity':'Pending','type':'Unmatched Transactions','count':unm,'details':'Transaction exists in only one source.'})
    oldc=conn.execute("SELECT COUNT(DISTINCT customer_code) c FROM customer_ledger WHERE posting_date < date('now','-90 day') GROUP BY customer_code HAVING ABS(SUM(debit-credit))>0.01").fetchall()
    if oldc: items.append({'severity':'Pending','type':'Customer 90+ Outstanding','count':len(oldc),'details':'Corporate customers with old outstanding balance.'})
    oldv=conn.execute("SELECT COUNT(DISTINCT vendor_code) c FROM vendor_ledger WHERE tx_date < date('now','-90 day') GROUP BY vendor_code HAVING ABS(SUM(debit-credit))>0.01").fetchall()
    if oldv: items.append({'severity':'Pending','type':'Vendor 90+ Outstanding','count':len(oldv),'details':'Vendors with old outstanding balance.'})
    conn.close(); return jsonify(items)

@app.post('/api/accounts/bank-reconcile')
@require_permission('ledger_view')
def bank_reconcile():
    bank=request.files.get('bank_file'); ledger=request.files.get('ledger_file')
    if not bank or not ledger: return jsonify({'error':'Select Bank Statement and Head Cash/SAP Ledger files'}),400
    a=_extract_doc_amount(_excel_rows(bank)); b=_extract_doc_amount(_excel_rows(ledger)); conn=db(); now=datetime.now().isoformat(timespec='seconds')
    cur=conn.execute("INSERT INTO reconciliation_runs(run_type,title,created_by,created_at) VALUES(?,?,?,?)",('Bank','Bank Statement vs Ledger',g.user.get('username'),now)); run_id=cur.lastrowid
    counts={'Matched':0,'Amount Difference':0,'Unmatched':0}
    for doc in sorted(set(a)|set(b)):
        av=float(a.get(doc,0)); bv=float(b.get(doc,0)); diff=round(av-bv,2)
        status='Matched' if doc in a and doc in b and abs(diff)<.01 else ('Amount Difference' if doc in a and doc in b else 'Unmatched'); counts[status]+=1
        conn.execute("INSERT INTO reconciliation_results(run_id,source_name,document_number,source_amount,target_amount,difference,status,details) VALUES(?,?,?,?,?,?,?,?)",(run_id,'Bank vs Ledger',doc,av,bv,diff,status,''))
    conn.commit(); conn.close(); audit('Create','Bank Reconciliation',f'Run {run_id}: {counts}')
    return jsonify({'ok':True,'run_id':run_id,'counts':counts})


@app.get('/api/accounts/daily-closing')
@require_permission('ledger_view')
def accounts_daily_closing():
    date_from=(request.args.get('from') or '').strip()
    date_to=(request.args.get('to') or '').strip()
    conn=db()
    where=[]; params=[]
    if date_from: where.append("posting_date>=?"); params.append(date_from)
    if date_to: where.append("posting_date<=?"); params.append(date_to)
    w=(' WHERE '+' AND '.join(where)) if where else ''
    cash=conn.execute(f"SELECT posting_date dt,cash_type,SUM(debit) debit,SUM(credit) credit FROM cash_ledger{w} GROUP BY posting_date,cash_type",params).fetchall()
    cwhere=[]; cparams=[]
    if date_from: cwhere.append("closing_date>=?"); cparams.append(date_from)
    if date_to: cwhere.append("closing_date<=?"); cparams.append(date_to)
    cw=(' WHERE '+' AND '.join(cwhere)) if cwhere else ''
    closings=conn.execute(f"SELECT closing_date dt,SUM(total_closing_cash) closing_cash,SUM(system_total_sale) system_sale,SUM(collection_difference) cash_difference,SUM(settlement_bank) bank_settlement,SUM(card_difference) card_difference FROM cashier_closings{cw} GROUP BY closing_date",cparams).fetchall()
    by={}
    def rec(dt):
        return by.setdefault(dt,{'date':dt,'head_debit':0,'head_credit':0,'petty_debit':0,'petty_credit':0,'closing_cash':0,'system_sale':0,'cash_difference':0,'bank_settlement':0,'card_difference':0})
    for r in cash:
        x=rec(r['dt'] or '')
        prefix='head' if str(r['cash_type']).lower().startswith('head') else 'petty'
        x[prefix+'_debit']+=float(r['debit'] or 0); x[prefix+'_credit']+=float(r['credit'] or 0)
    for r in closings:
        x=rec(r['dt'] or '')
        for k in ('closing_cash','system_sale','cash_difference','bank_settlement','card_difference'): x[k]=float(r[k] or 0)
    out=[]
    for x in by.values():
        x['head_balance']=round(x['head_debit']-x['head_credit'],2)
        x['petty_balance']=round(x['petty_debit']-x['petty_credit'],2)
        dif=abs(x['cash_difference'])+abs(x['card_difference'])
        x['status']='Complete' if dif<0.01 and x['system_sale'] else ('Difference' if dif>=0.01 else 'Pending')
        out.append(x)
    conn.close()
    return jsonify(sorted(out,key=lambda x:x['date'],reverse=True))

@app.get('/api/accounts/period-locks')
@require_permission('ledger_view')
def list_period_locks():
    conn=db(); rows=conn.execute("SELECT * FROM period_locks ORDER BY period_key DESC").fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.post('/api/accounts/period-locks')
@require_permission('documents_data_edit')
def add_period_lock():
    data=request.get_json(silent=True) or {}; key=str(data.get('period_key') or '').strip(); reason=str(data.get('reason') or '').strip()
    if not key or len(key)!=7: return jsonify({'error':'Period YYYY-MM format mein enter karein'}),400
    conn=db(); now=datetime.now().isoformat(timespec='seconds')
    conn.execute("INSERT OR REPLACE INTO period_locks(period_key,locked_by,locked_at,reason) VALUES(?,?,?,?)",(key,g.user.get('username'),now,reason)); conn.commit(); conn.close()
    audit('Lock','Period Lock',f'{key}: {reason}'); return jsonify({'ok':True})

@app.delete('/api/accounts/period-locks/<period_key>')
@require_permission('documents_data_edit')
def delete_period_lock(period_key):
    conn=db(); conn.execute("DELETE FROM period_locks WHERE period_key=?",(period_key,)); conn.commit(); conn.close()
    audit('Unlock','Period Lock',period_key); return jsonify({'ok':True})

@app.get('/api/accounts/approvals')
@require_permission('ledger_view')
def list_approvals():
    conn=db(); rows=conn.execute("SELECT * FROM approval_requests ORDER BY id DESC LIMIT 500").fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.post('/api/accounts/approvals')
@require_permission('ledger_edit')
def create_approval():
    data=request.get_json(silent=True) or {}; ref=str(data.get('reference_no') or '').strip()
    if not ref: return jsonify({'error':'Reference No. required'}),400
    now=datetime.now().isoformat(timespec='seconds'); conn=db()
    cur=conn.execute("INSERT INTO approval_requests(request_type,reference_no,amount,reason,maker,status,created_at,updated_at) VALUES(?,?,?,?,?,'Pending Checker',?,?)",(data.get('request_type') or 'Payment',ref,_num(data.get('amount')),data.get('reason') or '',g.user.get('username'),now,now)); conn.commit(); rid=cur.lastrowid; conn.close()
    audit('Create','Approval Workflow',f'Request {rid} / {ref}'); return jsonify({'ok':True,'id':rid})

@app.post('/api/accounts/approvals/<int:request_id>/action')
@require_permission('ledger_edit')
def approval_action(request_id):
    data=request.get_json(silent=True) or {}; action=str(data.get('action') or '').lower(); now=datetime.now().isoformat(timespec='seconds'); user=g.user.get('username')
    conn=db(); row=conn.execute("SELECT * FROM approval_requests WHERE id=?",(request_id,)).fetchone()
    if not row: conn.close(); return jsonify({'error':'Request not found'}),404
    if action=='check': conn.execute("UPDATE approval_requests SET checker=?,status='Pending Approver',updated_at=? WHERE id=?",(user,now,request_id))
    elif action=='approve': conn.execute("UPDATE approval_requests SET approver=?,status='Approved',updated_at=? WHERE id=?",(user,now,request_id))
    elif action=='reject': conn.execute("UPDATE approval_requests SET approver=?,status='Rejected',updated_at=? WHERE id=?",(user,now,request_id))
    else: conn.close(); return jsonify({'error':'Invalid action'}),400
    conn.commit(); conn.close(); audit(action.title(),'Approval Workflow',f'Request {request_id}'); return jsonify({'ok':True})

@app.get('/api/accounts/audit-report.xlsx')
@require_permission('reports_view')
def accounts_audit_report_excel():
    conn=db(); wb=Workbook(); ws=wb.active; ws.title='Audit Summary'
    ws.append(['Accounts Control & Audit Report']); ws.append(['Generated',datetime.now().strftime('%Y-%m-%d %H:%M:%S')]); ws.append([])
    ws.append(['Exception','Count','Amount'])
    dup=conn.execute("SELECT COUNT(*) c FROM duplicate_documents").fetchone()['c']; ws.append(['Duplicate Documents',dup,0])
    for status in ('Amount Difference','Unmatched','Matched'):
        r=conn.execute("SELECT COUNT(*) c,COALESCE(SUM(ABS(difference)),0) amount FROM reconciliation_results WHERE status=?",(status,)).fetchone(); ws.append([status,r['c'],r['amount']])
    wa=wb.create_sheet('Approvals'); wa.append(['ID','Type','Reference','Amount','Reason','Maker','Checker','Approver','Status','Created'])
    for r in conn.execute("SELECT * FROM approval_requests ORDER BY id DESC").fetchall(): wa.append([r['id'],r['request_type'],r['reference_no'],r['amount'],r['reason'],r['maker'],r['checker'],r['approver'],r['status'],r['created_at']])
    wl=wb.create_sheet('Period Locks'); wl.append(['Period','Locked By','Locked At','Reason'])
    for r in conn.execute("SELECT * FROM period_locks ORDER BY period_key DESC").fetchall(): wl.append([r['period_key'],r['locked_by'],r['locked_at'],r['reason']])
    conn.close()
    for sh in wb.worksheets:
        sh.freeze_panes='A2'
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width=min(40,max(12,max(len(str(c.value or '')) for c in col)+2))
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name=f"Accounts_Audit_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get('/api/accounts/cash-flow-forecast')
@require_permission('reports_view')
def cash_flow_forecast():
    days=max(7,min(180,int(request.args.get('days') or 30)))
    conn=db(); today=datetime.now().date(); start=(today-timedelta(days=days)).isoformat()
    cash=conn.execute("SELECT posting_date dt,SUM(debit-credit) net FROM cash_ledger WHERE posting_date>=? GROUP BY posting_date ORDER BY posting_date",(start,)).fetchall()
    cust=conn.execute("SELECT posting_date dt,SUM(debit-credit) net FROM customer_ledger WHERE posting_date>=? GROUP BY posting_date",(start,)).fetchall()
    vend=conn.execute("SELECT tx_date dt,SUM(credit-debit) net FROM vendor_ledger WHERE tx_date>=? GROUP BY tx_date",(start,)).fetchall()
    conn.close()
    hist={}
    for rows,key in ((cash,'cash'),(cust,'receivable'),(vend,'payable')):
        for r in rows:
            d=str(r['dt'] or '')[:10]
            if d: hist.setdefault(d,{'cash':0,'receivable':0,'payable':0})[key]+=float(r['net'] or 0)
    vals=list(hist.values())[-days:]
    avg_cash=sum(x['cash'] for x in vals)/max(1,len(vals)); avg_in=sum(max(0,x['receivable']) for x in vals)/max(1,len(vals)); avg_out=sum(max(0,x['payable']) for x in vals)/max(1,len(vals))
    balance=0.0; out=[]
    for i in range(1,days+1):
        d=today+timedelta(days=i); opening=balance; inflow=avg_cash+avg_in; outflow=avg_out; balance=opening+inflow-outflow
        out.append({'date':d.isoformat(),'opening':round(opening,2),'projected_inflow':round(inflow,2),'projected_outflow':round(outflow,2),'closing':round(balance,2),'risk':'Low Cash' if balance<0 else 'Normal'})
    return jsonify(out)

@app.get('/api/accounts/budget-vs-actual')
@require_permission('reports_view')
def budget_vs_actual():
    period=(request.args.get('period') or datetime.now().strftime('%Y-%m')).strip(); conn=db()
    budgets=conn.execute("SELECT * FROM finance_budgets WHERE period_key=? ORDER BY account_head",(period,)).fetchall()
    actuals={
      'Head Cash': conn.execute("SELECT COALESCE(SUM(debit-credit),0) v FROM cash_ledger WHERE cash_type='Head Cash' AND substr(posting_date,1,7)=?",(period,)).fetchone()['v'],
      'Petty Cash': conn.execute("SELECT COALESCE(SUM(debit-credit),0) v FROM cash_ledger WHERE cash_type='Petty Cash' AND substr(posting_date,1,7)=?",(period,)).fetchone()['v'],
      'Corporate Receivable': conn.execute("SELECT COALESCE(SUM(debit-credit),0) v FROM customer_ledger WHERE substr(posting_date,1,7)=?",(period,)).fetchone()['v'],
      'Vendor Payable': conn.execute("SELECT COALESCE(SUM(credit-debit),0) v FROM vendor_ledger WHERE substr(tx_date,1,7)=?",(period,)).fetchone()['v']
    }; conn.close()
    return jsonify([dict(r)|{'actual_amount':round(float(actuals.get(r['account_head'],0) or 0),2),'variance':round(float(r['budget_amount'] or 0)-float(actuals.get(r['account_head'],0) or 0),2)} for r in budgets])

@app.post('/api/accounts/budgets')
@require_permission('ledger_edit')
def save_budget():
    d=request.get_json(silent=True) or {}; period=str(d.get('period_key') or '').strip(); head=str(d.get('account_head') or '').strip()
    if len(period)!=7 or not head: return jsonify({'error':'Period and account head required'}),400
    conn=db(); now=datetime.now().isoformat(timespec='seconds')
    conn.execute("INSERT INTO finance_budgets(period_key,cost_center,account_head,budget_amount,created_by,created_at) VALUES(?,?,?,?,?,?) ON CONFLICT(period_key,cost_center,account_head) DO UPDATE SET budget_amount=excluded.budget_amount,created_by=excluded.created_by,created_at=excluded.created_at",(period,str(d.get('cost_center') or 'ALL'),head,_num(d.get('budget_amount')),g.user.get('username'),now)); conn.commit(); conn.close()
    audit('Save','Budget Control',f'{period} / {head}'); return jsonify({'ok':True})

@app.get('/api/accounts/month-end-tasks')
@require_permission('reports_view')
def month_end_tasks():
    period=(request.args.get('period') or datetime.now().strftime('%Y-%m')).strip(); conn=db()
    defaults=['Bank reconciliation complete','Head cash verified','Petty cash verified','Corporate aging reviewed','Vendor aging reviewed','Duplicate documents cleared','Approval requests closed','Audit report generated','Period locked']
    now=datetime.now().isoformat(timespec='seconds')
    for task in defaults:
        conn.execute("INSERT OR IGNORE INTO month_end_tasks(period_key,task_name,status,updated_at) VALUES(?,?,'Pending',?)",(period,task,now))
    conn.commit(); rows=conn.execute("SELECT * FROM month_end_tasks WHERE period_key=? ORDER BY id",(period,)).fetchall(); conn.close(); return jsonify([dict(r) for r in rows])

@app.post('/api/accounts/month-end-tasks/<int:task_id>')
@require_permission('ledger_edit')
def update_month_end_task(task_id):
    d=request.get_json(silent=True) or {}; status=str(d.get('status') or 'Pending'); remarks=str(d.get('remarks') or '')
    conn=db(); conn.execute("UPDATE month_end_tasks SET status=?,remarks=?,updated_by=?,updated_at=? WHERE id=?",(status,remarks,g.user.get('username'),datetime.now().isoformat(timespec='seconds'),task_id)); conn.commit(); conn.close()
    audit('Update','Month End Close',f'Task {task_id}: {status}'); return jsonify({'ok':True})



@app.get('/api/accounts/journal-vouchers')
@require_permission('reports_view')
def journal_vouchers_list():
    period=(request.args.get('period') or '').strip()
    conn=db(); params=[]; where=''
    if period:
        where='WHERE substr(v.posting_date,1,7)=?'; params=[period]
    rows=conn.execute(f"""SELECT v.id,v.voucher_no,v.posting_date,v.narration,v.status,v.created_by,v.created_at,
        ROUND(COALESCE(SUM(l.debit),0),2) total_debit,ROUND(COALESCE(SUM(l.credit),0),2) total_credit,
        COUNT(l.id) line_count FROM journal_vouchers v LEFT JOIN journal_lines l ON l.voucher_id=v.id
        {where} GROUP BY v.id ORDER BY v.posting_date DESC,v.id DESC LIMIT 500""",params).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.post('/api/accounts/journal-vouchers')
@require_permission('ledger_edit')
def create_journal_voucher():
    d=request.get_json(silent=True) or {}; lines=d.get('lines') or []
    posting_date=str(d.get('posting_date') or '').strip(); narration=str(d.get('narration') or '').strip()
    clean=[]
    for line in lines:
        account=str(line.get('account_name') or '').strip(); debit=_num(line.get('debit')); credit=_num(line.get('credit'))
        if account and (debit>0 or credit>0): clean.append((str(line.get('account_code') or ''),account,str(line.get('cost_center') or 'ALL'),debit,credit))
    td=round(sum(x[3] for x in clean),2); tc=round(sum(x[4] for x in clean),2)
    if not posting_date or len(clean)<2: return jsonify({'error':'Posting date and minimum two journal lines required'}),400
    if td<=0 or abs(td-tc)>0.01: return jsonify({'error':f'Journal not balanced. Debit {td:,.2f}, Credit {tc:,.2f}'}),400
    conn=db()
    if conn.execute("SELECT 1 FROM period_locks WHERE period_key=?",(posting_date[:7],)).fetchone(): conn.close(); return jsonify({'error':'Selected accounting period is locked'}),409
    prefix='JV-'+posting_date.replace('-','')+'-'; seq=conn.execute("SELECT COUNT(*) c FROM journal_vouchers WHERE posting_date=?",(posting_date,)).fetchone()['c']+1; voucher_no=f'{prefix}{seq:04d}'
    now=datetime.now().isoformat(timespec='seconds'); cur=conn.execute("INSERT INTO journal_vouchers(voucher_no,posting_date,narration,status,created_by,created_at) VALUES(?,?,?,?,?,?)",(voucher_no,posting_date,narration,'Posted',g.user.get('username'),now)); vid=cur.lastrowid
    conn.executemany("INSERT INTO journal_lines(voucher_id,account_code,account_name,cost_center,debit,credit) VALUES(?,?,?,?,?,?)",[(vid,*x) for x in clean]); conn.commit(); conn.close()
    audit('Post','Journal Voucher',f'{voucher_no} / {td:,.2f}'); return jsonify({'ok':True,'voucher_no':voucher_no})

@app.get('/api/accounts/trial-balance')
@require_permission('reports_view')
def finance_trial_balance():
    period=(request.args.get('period') or datetime.now().strftime('%Y-%m')).strip(); conn=db(); data={}
    def add(name,debit=0,credit=0):
        r=data.setdefault(name,{'account_name':name,'debit':0.0,'credit':0.0}); r['debit']+=float(debit or 0); r['credit']+=float(credit or 0)
    for r in conn.execute("SELECT cash_type account_name,SUM(debit) debit,SUM(credit) credit FROM cash_ledger WHERE substr(posting_date,1,7)=? GROUP BY cash_type",(period,)).fetchall(): add(r['account_name'],r['debit'],r['credit'])
    r=conn.execute("SELECT SUM(debit) debit,SUM(credit) credit FROM customer_ledger WHERE substr(posting_date,1,7)=?",(period,)).fetchone(); add('Corporate Receivables',r['debit'],r['credit'])
    r=conn.execute("SELECT SUM(debit) debit,SUM(credit) credit FROM vendor_ledger WHERE substr(tx_date,1,7)=?",(period,)).fetchone(); add('Vendor Payables',r['debit'],r['credit'])
    for r in conn.execute("SELECT l.account_name,SUM(l.debit) debit,SUM(l.credit) credit FROM journal_lines l JOIN journal_vouchers v ON v.id=l.voucher_id WHERE substr(v.posting_date,1,7)=? AND v.status='Posted' GROUP BY l.account_name",(period,)).fetchall(): add(r['account_name'],r['debit'],r['credit'])
    conn.close(); rows=[]
    for x in sorted(data.values(),key=lambda z:z['account_name']): x['debit']=round(x['debit'],2); x['credit']=round(x['credit'],2); x['balance']=round(x['debit']-x['credit'],2); rows.append(x)
    return jsonify({'period':period,'rows':rows,'total_debit':round(sum(x['debit'] for x in rows),2),'total_credit':round(sum(x['credit'] for x in rows),2)})

@app.get('/api/accounts/accruals')
@require_permission('reports_view')
def accruals_list():
    conn=db(); rows=conn.execute("SELECT * FROM accrual_schedules ORDER BY start_date DESC,id DESC LIMIT 500").fetchall(); conn.close(); out=[]
    today=datetime.now().date()
    for r in rows:
        x=dict(r)
        try:
            s=datetime.fromisoformat(x['start_date']).date(); e=datetime.fromisoformat(x['end_date']).date(); total_months=max(1,(e.year-s.year)*12+e.month-s.month+1); elapsed=max(0,min(total_months,(today.year-s.year)*12+today.month-s.month+1)); auto=round(float(x['total_amount'] or 0)*elapsed/total_months,2)
        except Exception: auto=float(x['recognized_amount'] or 0)
        x['calculated_recognized']=auto; x['outstanding']=round(float(x['total_amount'] or 0)-auto,2); x['status']='Closed' if x['outstanding']<=0.01 else 'Open'; out.append(x)
    return jsonify(out)

@app.post('/api/accounts/accruals')
@require_permission('ledger_edit')
def create_accrual():
    d=request.get_json(silent=True) or {}; desc=str(d.get('description') or '').strip(); start=str(d.get('start_date') or ''); end=str(d.get('end_date') or '')
    if not desc or not start or not end or _num(d.get('total_amount'))<=0: return jsonify({'error':'Description, dates and positive amount required'}),400
    conn=db(); conn.execute("INSERT INTO accrual_schedules(schedule_type,reference_no,description,start_date,end_date,total_amount,account_name,cost_center,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",(str(d.get('schedule_type') or 'Accrual'),str(d.get('reference_no') or ''),desc,start,end,_num(d.get('total_amount')),str(d.get('account_name') or ''),str(d.get('cost_center') or 'ALL'),g.user.get('username'),datetime.now().isoformat(timespec='seconds'))); conn.commit(); conn.close(); audit('Create','Accrual Schedule',desc); return jsonify({'ok':True})

@app.get('/api/accounts/fixed-assets')
@require_permission('reports_view')
def fixed_assets_list():
    conn=db(); rows=conn.execute("SELECT * FROM fixed_assets ORDER BY acquisition_date DESC,id DESC LIMIT 1000").fetchall(); conn.close(); out=[]; today=datetime.now().date()
    for r in rows:
        x=dict(r); cost=float(x['cost'] or 0); salvage=float(x['salvage_value'] or 0); life=max(1,int(x['useful_life_months'] or 60))
        try: acq=datetime.fromisoformat(x['acquisition_date']).date(); used=max(0,min(life,(today.year-acq.year)*12+today.month-acq.month))
        except Exception: used=0
        monthly=(cost-salvage)/life; accum=min(cost-salvage,monthly*used); x['monthly_depreciation']=round(monthly,2); x['accumulated_depreciation']=round(accum,2); x['net_book_value']=round(cost-accum,2); out.append(x)
    return jsonify(out)

@app.post('/api/accounts/fixed-assets')
@require_permission('ledger_edit')
def create_fixed_asset():
    d=request.get_json(silent=True) or {}; code=str(d.get('asset_code') or '').strip(); name=str(d.get('asset_name') or '').strip(); date=str(d.get('acquisition_date') or '')
    if not code or not name or not date or _num(d.get('cost'))<=0: return jsonify({'error':'Asset code, name, acquisition date and cost required'}),400
    conn=db()
    try:
        conn.execute("INSERT INTO fixed_assets(asset_code,asset_name,category,acquisition_date,cost,salvage_value,useful_life_months,location,custodian,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(code,name,str(d.get('category') or ''),date,_num(d.get('cost')),_num(d.get('salvage_value')),max(1,int(d.get('useful_life_months') or 60)),str(d.get('location') or ''),str(d.get('custodian') or ''),g.user.get('username'),datetime.now().isoformat(timespec='seconds'))); conn.commit()
    except sqlite3.IntegrityError: conn.close(); return jsonify({'error':'Asset code already exists'}),409
    conn.close(); audit('Create','Fixed Asset',f'{code} - {name}'); return jsonify({'ok':True})


@app.get('/api/accounts/finance-health')
@require_permission('reports_view')
def finance_health():
    period=(request.args.get('period') or datetime.now().strftime('%Y-%m')).strip(); conn=db()
    def scalar(sql,args=()):
        row=conn.execute(sql,args).fetchone(); return float((row[0] if row else 0) or 0)
    customer=scalar("SELECT COALESCE(SUM(debit-credit),0) FROM customer_ledger WHERE substr(posting_date,1,7)<=?",(period,))
    vendor=scalar("SELECT COALESCE(SUM(credit-debit),0) FROM vendor_ledger WHERE substr(tx_date,1,7)<=?",(period,))
    cash=scalar("SELECT COALESCE(SUM(debit-credit),0) FROM cash_ledger WHERE substr(posting_date,1,7)<=?",(period,))
    exceptions=scalar("SELECT COUNT(*) FROM reconciliation_results WHERE status!='Matched'")
    pending=scalar("SELECT COUNT(*) FROM approval_requests WHERE status NOT IN ('Approved','Rejected')")
    overdue=scalar("SELECT COUNT(*) FROM payment_schedules WHERE due_date<date('now') AND status NOT IN ('Paid','Cancelled')")
    locked=bool(conn.execute("SELECT 1 FROM period_locks WHERE period_key=?",(period,)).fetchone())
    close_rows=conn.execute("SELECT status,COUNT(*) c FROM month_end_tasks WHERE period_key=? GROUP BY status",(period,)).fetchall(); close={r['status']:r['c'] for r in close_rows}
    conn.close(); total=sum(close.values()); complete=close.get('Completed',0)+close.get('Complete',0)
    score=100-min(35,int(exceptions)*2)-min(20,int(pending)*3)-min(20,int(overdue)*3)-(0 if locked else 10)-(0 if not total or complete==total else 15)
    return jsonify({'period':period,'health_score':max(0,score),'cash_position':round(cash,2),'customer_receivable':round(customer,2),'vendor_payable':round(vendor,2),'reconciliation_exceptions':int(exceptions),'pending_approvals':int(pending),'overdue_payments':int(overdue),'period_locked':locked,'close_completed':complete,'close_total':total})

@app.get('/api/accounts/financial-statements')
@require_permission('reports_view')
def financial_statements():
    period=(request.args.get('period') or datetime.now().strftime('%Y-%m')).strip(); conn=db()
    def sumq(sql):
        r=conn.execute(sql,(period,)).fetchone(); return float((r[0] if r else 0) or 0)
    cash=sumq("SELECT COALESCE(SUM(debit-credit),0) FROM cash_ledger WHERE substr(posting_date,1,7)<=?")
    receivable=sumq("SELECT COALESCE(SUM(debit-credit),0) FROM customer_ledger WHERE substr(posting_date,1,7)<=?")
    payable=sumq("SELECT COALESCE(SUM(credit-debit),0) FROM vendor_ledger WHERE substr(tx_date,1,7)<=?")
    assets=conn.execute("SELECT cost,salvage_value,useful_life_months,acquisition_date FROM fixed_assets WHERE substr(acquisition_date,1,7)<=? AND status='Active'",(period,)).fetchall()
    end=datetime.fromisoformat(period+'-01').date(); nbv=0
    for a in assets:
        try:
            d=datetime.fromisoformat(a['acquisition_date']).date(); life=max(1,int(a['useful_life_months'] or 60)); used=max(0,min(life,(end.year-d.year)*12+end.month-d.month+1)); cost=float(a['cost'] or 0); salvage=float(a['salvage_value'] or 0); nbv+=cost-min(cost-salvage,(cost-salvage)/life*used)
        except Exception: pass
    jv=conn.execute("SELECT l.account_name,SUM(l.debit) debit,SUM(l.credit) credit FROM journal_lines l JOIN journal_vouchers v ON v.id=l.voucher_id WHERE substr(v.posting_date,1,7)=? AND v.status='Posted' GROUP BY l.account_name",(period,)).fetchall()
    income=[]; expense=[]
    for r in jv:
        name=str(r['account_name'] or ''); net=float(r['credit'] or 0)-float(r['debit'] or 0); low=name.lower()
        if any(k in low for k in ('sale','income','revenue','gain')): income.append({'account':name,'amount':round(net,2)})
        elif any(k in low for k in ('expense','cost','salary','rent','utility','loss','depreciation')): expense.append({'account':name,'amount':round(-net,2)})
    total_income=sum(x['amount'] for x in income); total_expense=sum(x['amount'] for x in expense); profit=total_income-total_expense
    conn.close(); return jsonify({'period':period,'income':income,'expenses':expense,'total_income':round(total_income,2),'total_expense':round(total_expense,2),'net_profit':round(profit,2),'assets':[{'account':'Cash & Cash Equivalents','amount':round(cash,2)},{'account':'Corporate Receivables','amount':round(receivable,2)},{'account':'Fixed Assets - NBV','amount':round(nbv,2)}],'liabilities':[{'account':'Vendor Payables','amount':round(payable,2)}],'total_assets':round(cash+receivable+nbv,2),'total_liabilities':round(payable,2),'management_view':True})

@app.get('/api/accounts/cost-centers')
@require_permission('reports_view')
def cost_center_analysis():
    period=(request.args.get('period') or datetime.now().strftime('%Y-%m')).strip(); conn=db()
    rows=conn.execute("SELECT COALESCE(NULLIF(TRIM(l.cost_center),''),'ALL') cost_center,SUM(l.debit) debit,SUM(l.credit) credit,COUNT(*) entries FROM journal_lines l JOIN journal_vouchers v ON v.id=l.voucher_id WHERE substr(v.posting_date,1,7)=? AND v.status='Posted' GROUP BY 1 ORDER BY 1",(period,)).fetchall(); conn.close()
    out=[]
    for r in rows: out.append({'cost_center':r['cost_center'],'debit':round(float(r['debit'] or 0),2),'credit':round(float(r['credit'] or 0),2),'net':round(float(r['credit'] or 0)-float(r['debit'] or 0),2),'entries':r['entries']})
    return jsonify({'period':period,'rows':out})

@app.get('/api/accounts/tax-register')
@require_permission('reports_view')
def tax_register_list():
    period=(request.args.get('period') or '').strip(); conn=db(); where=''; args=[]
    if period: where='WHERE substr(posting_date,1,7)=?'; args=[period]
    rows=conn.execute(f"SELECT * FROM tax_register {where} ORDER BY posting_date DESC,id DESC LIMIT 1000",args).fetchall(); conn.close(); return jsonify([dict(r) for r in rows])

@app.post('/api/accounts/tax-register')
@require_permission('ledger_edit')
def tax_register_create():
    d=request.get_json(silent=True) or {}; date=str(d.get('posting_date') or ''); taxable=_num(d.get('taxable_amount')); rate=_num(d.get('tax_rate')); amount=_num(d.get('tax_amount')) or round(taxable*rate/100,2)
    if not date or not str(d.get('tax_type') or '').strip(): return jsonify({'error':'Posting date and tax type required'}),400
    conn=db(); conn.execute("INSERT INTO tax_register(posting_date,tax_type,party_type,party_code,party_name,document_number,taxable_amount,tax_rate,tax_amount,status,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(date,str(d.get('tax_type')),str(d.get('party_type') or 'Vendor'),str(d.get('party_code') or ''),str(d.get('party_name') or ''),str(d.get('document_number') or ''),taxable,rate,amount,str(d.get('status') or 'Pending'),g.user.get('username'),datetime.now().isoformat(timespec='seconds'))); conn.commit(); conn.close(); audit('Create','Tax Register',str(d.get('document_number') or d.get('tax_type'))); return jsonify({'ok':True})

@app.get('/api/accounts/payment-calendar')
@require_permission('reports_view')
def payment_calendar_list():
    conn=db(); rows=conn.execute("SELECT *,CAST(julianday(due_date)-julianday(date('now')) AS INTEGER) days_to_due FROM payment_schedules ORDER BY CASE status WHEN 'Paid' THEN 2 ELSE 1 END,due_date,id LIMIT 1000").fetchall(); conn.close(); return jsonify([dict(r) for r in rows])

@app.post('/api/accounts/payment-calendar')
@require_permission('ledger_edit')
def payment_calendar_create():
    d=request.get_json(silent=True) or {}; due=str(d.get('due_date') or ''); name=str(d.get('party_name') or '').strip(); amount=_num(d.get('amount'))
    if not due or not name or amount<=0: return jsonify({'error':'Due date, party name and positive amount required'}),400
    conn=db(); conn.execute("INSERT INTO payment_schedules(due_date,party_type,party_code,party_name,reference_no,amount,priority,payment_method,status,remarks,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(due,str(d.get('party_type') or 'Vendor'),str(d.get('party_code') or ''),name,str(d.get('reference_no') or ''),amount,str(d.get('priority') or 'Normal'),str(d.get('payment_method') or ''),str(d.get('status') or 'Planned'),str(d.get('remarks') or ''),g.user.get('username'),datetime.now().isoformat(timespec='seconds'))); conn.commit(); conn.close(); audit('Create','Payment Calendar',f'{name}: {amount:,.2f}'); return jsonify({'ok':True})

@app.post('/api/accounts/payment-calendar/<int:item_id>/status')
@require_permission('ledger_edit')
def payment_calendar_status(item_id):
    d=request.get_json(silent=True) or {}; status=str(d.get('status') or 'Planned'); conn=db(); conn.execute("UPDATE payment_schedules SET status=? WHERE id=?",(status,item_id)); conn.commit(); conn.close(); audit('Update','Payment Calendar',f'{item_id}: {status}'); return jsonify({'ok':True})

# Initialize the database whenever the module is imported. This is required for
# production servers such as Gunicorn/Render, which load the Flask application
# with ``app:app`` and do not execute the __main__ block.
try:
    init_db()
    initialize_store_databases()
    create_automatic_backup()
except Exception as startup_error:
    # Keep a clear startup message in hosting logs instead of allowing login to
    # fail later with an unclear "no such table: users" error.
    print(f"Database startup initialization failed: {startup_error}")
    raise

if __name__=="__main__":
    print(f"Permanent data location: {DB}")
    print("\nCorporate Customer + Scrap Vendor Management System")
    print("Computer: http://127.0.0.1:5055")
    app.run(host="0.0.0.0",port=5055,debug=False)
